from fastapi import FastAPI, Depends, HTTPException, Query, File, UploadFile, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from starlette.middleware.base import BaseHTTPMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timedelta
from contextvars import ContextVar
import statistics
import secrets
import os
import base64
import json
import re
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from urllib.parse import quote

# .env 파일 자동 로드 (패키지 불필요)
def _load_dotenv():
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if not os.path.exists(env_path):
        return
    with open(env_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            key = key.strip(); val = val.strip()
            if key and val and key not in os.environ:
                os.environ[key] = val
_load_dotenv()

from database import (
    get_db, create_tables, Material, Supplier, Recipe,
    RecipeIngredient, Receipt, Device, Production,
    SemiProduct, FinishedProduct, Process, ProductBOM,
    ProductProcess, MaterialSupplier, ProductionPlan,
    SalesLog, SemiProductBOM, SemiProductProcess, StockAdjustment,
    Shipment, ReceiptNonconformance,
    TenantUser, UserSession, hash_password, verify_password, ensure_admin,
    SessionLocal,
)

# ─────────────────────────────────────────
# 인증 컨텍스트 (ContextVar — asyncio task-safe)
# ─────────────────────────────────────────
_uid_ctx: ContextVar[int] = ContextVar('uid_ctx', default=0)
_role_ctx: ContextVar[str] = ContextVar('role_ctx', default='user')

def uid() -> int:
    return _uid_ctx.get()

def current_role() -> str:
    return _role_ctx.get()

def require_admin():
    if current_role() != 'admin':
        raise HTTPException(403, "관리자 권한이 필요합니다")

# 인증 불필요 경로
_EXEMPT = {'/api/auth/login', '/api/auth/admin/login', '/api/equipment/production', '/'}

class _AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        if path in _EXEMPT or not path.startswith('/api/'):
            return await call_next(request)

        token = request.headers.get('X-Token', '')
        if not token:
            return JSONResponse({'detail': 'Not authenticated'}, status_code=401)

        db = SessionLocal()
        try:
            sess = db.query(UserSession).filter(
                UserSession.token == token,
                UserSession.expires_at > datetime.utcnow()
            ).first()
            if not sess:
                return JSONResponse({'detail': 'Session expired'}, status_code=401)
            t1 = _uid_ctx.set(sess.user_id)
            t2 = _role_ctx.set(sess.role)
            try:
                return await call_next(request)
            finally:
                _uid_ctx.reset(t1)
                _role_ctx.reset(t2)
        finally:
            db.close()

app = FastAPI(title="Foodly API", version="1.0.0")

app.add_middleware(_AuthMiddleware)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# HTML 서빙
@app.get("/", response_class=FileResponse)
def serve_html():
    return FileResponse("foodly.html", headers={"Cache-Control": "no-store"})


@app.on_event("startup")
def startup():
    create_tables()
    from seed import seed
    seed()
    db = SessionLocal()
    try:
        ensure_admin(db)
    finally:
        db.close()


# ─────────────────────────────────────────
# 인증 엔드포인트
# ─────────────────────────────────────────
class LoginIn(BaseModel):
    username: str
    business_number: str
    password: str

class AdminLoginIn(BaseModel):
    username: str
    password: str

class UserCreateIn(BaseModel):
    username: str
    password: str
    business_number: str
    company_name: str
    contact_person: Optional[str] = None
    contact: Optional[str] = None

class UserUpdateIn(BaseModel):
    password: Optional[str] = None
    business_number: Optional[str] = None
    company_name: Optional[str] = None
    contact_person: Optional[str] = None
    contact: Optional[str] = None
    status: Optional[str] = None

_SESSION_HOURS = 8

def _make_session(db, user_id: int, role: str) -> str:
    token = secrets.token_hex(32)
    db.add(UserSession(
        user_id=user_id,
        role=role,
        token=token,
        expires_at=datetime.utcnow() + timedelta(hours=_SESSION_HOURS),
    ))
    db.commit()
    return token

@app.post("/api/auth/login")
def user_login(data: LoginIn, db: Session = Depends(get_db)):
    user = db.query(TenantUser).filter(
        TenantUser.username == data.username,
        TenantUser.status == 'active',
        TenantUser.role == 'user',
    ).first()
    if not user or user.business_number != data.business_number or not verify_password(data.password, user.password_hash):
        raise HTTPException(401, "아이디, 사업자등록번호 또는 비밀번호가 올바르지 않습니다")
    token = _make_session(db, user.id, 'user')
    return {"token": token, "role": "user", "company_name": user.company_name, "username": user.username}

@app.post("/api/auth/admin/login")
def admin_login(data: AdminLoginIn, db: Session = Depends(get_db)):
    admin = db.query(TenantUser).filter(
        TenantUser.username == data.username,
        TenantUser.role == 'admin',
        TenantUser.status == 'active',
    ).first()
    if not admin or not verify_password(data.password, admin.password_hash):
        raise HTTPException(401, "관리자 아이디 또는 비밀번호가 올바르지 않습니다")
    token = _make_session(db, admin.id, 'admin')
    return {"token": token, "role": "admin", "username": admin.username}

@app.post("/api/auth/logout")
def logout(db: Session = Depends(get_db)):
    # 미들웨어에서 이미 토큰 검증됨; 세션 삭제
    # (요청 헤더에서 토큰 재추출)
    return {"ok": True}

@app.get("/api/auth/me")
def me(db: Session = Depends(get_db)):
    if current_role() == 'admin':
        return {"role": "admin", "username": "paxc", "company_name": "PAXC 운영사"}
    user = db.query(TenantUser).get(uid())
    if not user:
        raise HTTPException(404, "User not found")
    return {"role": "user", "username": user.username, "company_name": user.company_name}


@app.get("/api/auth/company")
def get_company(db: Session = Depends(get_db)):
    if current_role() == 'admin':
        return {"company_name": "PAXC 운영사", "business_number": "000-00-00000", "contact": "", "seal_image": None}
    user = db.query(TenantUser).get(uid())
    if not user:
        raise HTTPException(404, "User not found")
    return {
        "company_name": user.company_name,
        "business_number": user.business_number,
        "contact": user.contact or "",
        "seal_image": user.seal_image,
    }


class SealUpdate(BaseModel):
    seal_image: Optional[str] = None   # base64 data URL, None이면 삭제


@app.put("/api/auth/company/seal")
def update_seal(data: SealUpdate, db: Session = Depends(get_db)):
    if current_role() == 'admin':
        raise HTTPException(403, "관리자 계정은 직인을 등록할 수 없습니다")
    user = db.query(TenantUser).get(uid())
    if not user:
        raise HTTPException(404, "User not found")
    user.seal_image = data.seal_image
    db.commit()
    return {"ok": True}


# ─────────────────────────────────────────
# 관리자 전용: 사용자 관리
# ─────────────────────────────────────────
@app.get("/api/admin/users")
def admin_list_users(db: Session = Depends(get_db)):
    require_admin()
    users = db.query(TenantUser).filter(TenantUser.role == 'user').order_by(TenantUser.created_at).all()
    return [
        {"id": u.id, "username": u.username, "company_name": u.company_name,
         "business_number": u.business_number, "contact_person": u.contact_person,
         "contact": u.contact, "status": u.status,
         "created_at": u.created_at.strftime("%Y-%m-%d") if u.created_at else ""}
        for u in users
    ]

@app.post("/api/admin/users", status_code=201)
def admin_create_user(data: UserCreateIn, db: Session = Depends(get_db)):
    require_admin()
    if db.query(TenantUser).filter(TenantUser.username == data.username).first():
        raise HTTPException(409, "이미 존재하는 아이디입니다")
    u = TenantUser(
        username=data.username,
        password_hash=hash_password(data.password),
        business_number=data.business_number,
        company_name=data.company_name,
        contact_person=data.contact_person,
        contact=data.contact,
        role='user',
        status='active',
    )
    db.add(u); db.commit(); db.refresh(u)
    return {"id": u.id, "username": u.username}

@app.put("/api/admin/users/{user_id}")
def admin_update_user(user_id: int, data: UserUpdateIn, db: Session = Depends(get_db)):
    require_admin()
    u = db.query(TenantUser).filter(TenantUser.id == user_id, TenantUser.role == 'user').first()
    if not u:
        raise HTTPException(404, "User not found")
    if data.password:
        u.password_hash = hash_password(data.password)
    if data.business_number is not None: u.business_number = data.business_number
    if data.company_name is not None:    u.company_name = data.company_name
    if data.contact_person is not None:  u.contact_person = data.contact_person
    if data.contact is not None:         u.contact = data.contact
    if data.status is not None:          u.status = data.status
    db.commit()
    return {"ok": True}

@app.delete("/api/admin/users/{user_id}")
def admin_delete_user(user_id: int, db: Session = Depends(get_db)):
    require_admin()
    u = db.query(TenantUser).filter(TenantUser.id == user_id, TenantUser.role == 'user').first()
    if not u:
        raise HTTPException(404, "User not found")
    db.delete(u); db.commit()
    return {"ok": True}


# ─────────────────────────────────────────
# Pydantic Schemas
# ─────────────────────────────────────────
class MaterialCreate(BaseModel):
    name: str
    unit: str = "kg"
    safety_stock: float = 0

class MaterialUpdate(BaseModel):
    name: Optional[str] = None
    unit: Optional[str] = None
    safety_stock: Optional[float] = None
    current_stock: Optional[float] = None

class SupplierCreate(BaseModel):
    name: str
    business_number: Optional[str] = None
    contact: Optional[str] = None

class RecipeCreate(BaseModel):
    product_code: str
    product_name: str
    category: Optional[str] = None
    base_quantity: float = 1
    base_unit: str = "개"
    version: str = "v1.0"
    status: str = "active"
    note: Optional[str] = None

class RecipeIngredientCreate(BaseModel):
    material_id: int
    quantity: float
    unit: str = "kg"

class ReceiptCreate(BaseModel):
    material_id: int
    supplier_id: Optional[int] = None
    lot_number: Optional[str] = None
    quantity: float
    unit_price: Optional[float] = None
    delivery_date: Optional[str] = None
    expiry_date: Optional[str] = None
    input_method: str = "manual"
    packaging_ok: Optional[bool] = None
    visual_ok: Optional[bool] = None
    judgment_ok: Optional[bool] = None
    inspector: Optional[str] = None
    confirmer: Optional[str] = None

class NonconformanceCreate(BaseModel):
    receipt_id: int
    date: Optional[str] = None
    content: str
    inspector: Optional[str] = None

class ProductionCreate(BaseModel):
    lot_number: Optional[str] = None
    finished_product_id: Optional[int] = None
    recipe_id: Optional[int] = None
    device_id: Optional[int] = None
    produced_quantity: Optional[float] = None
    good_quantity: Optional[float] = None
    defect_quantity: float = 0
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    input_method: str = "manual"
    status: str = "completed"
    note: Optional[str] = None

class ProductionPlanCreate(BaseModel):
    product_id: int
    planned_date: str
    planned_quantity: float
    note: Optional[str] = None
    status: str = "planned"

class DeviceProductionIn(BaseModel):
    device_id: str
    product_code: str
    produced_quantity: float
    defects: Optional[List[dict]] = []

class ShipmentCreate(BaseModel):
    finished_product_id: int
    customer_id: Optional[int] = None
    production_id: Optional[int] = None
    lot_number: Optional[str] = None
    quantity: float
    unit_price: Optional[float] = None
    delivery_date: Optional[str] = None
    note: Optional[str] = None

class ShipmentUpdate(BaseModel):
    finished_product_id: Optional[int] = None
    customer_id: Optional[int] = None
    lot_number: Optional[str] = None
    quantity: Optional[float] = None
    unit_price: Optional[float] = None
    delivery_date: Optional[str] = None
    note: Optional[str] = None


# ─────────────────────────────────────────
# BOM 재귀 전개 헬퍼
# ─────────────────────────────────────────
def _to_base_kg(qty: float, unit: str) -> float:
    if unit in ('g', 'ml'): return qty / 1000
    return qty  # kg, L → 1:1

def _conv_eff(obj, qty: float) -> float:
    """단위환산 적용 후 kg 환산값 (% BOM 계산용) — FinishedProduct/SemiProduct 공용"""
    if not obj or not obj.unit_conv_qty or not obj.unit_conv_unit:
        return qty
    u1 = obj.unit_conv_unit
    if u1 in ('kg', 'g', 'L', 'ml'):
        return qty * _to_base_kg(obj.unit_conv_qty, u1)
    if u1 == 'ea' and obj.unit_conv2_qty and obj.unit_conv2_unit:
        return qty * obj.unit_conv_qty * _to_base_kg(obj.unit_conv2_qty, obj.unit_conv2_unit)
    return qty

def explode_semi_bom(semi_id: int, qty: float, db: Session,
                     visited: frozenset = frozenset()) -> dict:
    """
    반제품 BOM을 재귀 전개해 원재료별 차감량 반환.
    sub_semi_id(반제품 구성) 항목도 재귀 전개.
    반환: {material_id: deduct_amount}
    """
    if semi_id in visited:
        return {}
    visited = visited | {semi_id}

    semi = db.query(SemiProduct).get(semi_id)
    if not semi:
        return {}

    eff = _conv_eff(semi, qty)
    result: dict = {}

    for bom in db.query(SemiProductBOM).filter(SemiProductBOM.semi_product_id == semi_id).all():
        if bom.material_id:
            amt = eff * (bom.quantity / 100) if bom.unit == '%' else bom.quantity * qty
            result[bom.material_id] = result.get(bom.material_id, 0) + amt
        elif bom.sub_semi_id:
            child_qty = eff * (bom.quantity / 100) if bom.unit == '%' else bom.quantity * qty
            for mat_id, amt in explode_semi_bom(bom.sub_semi_id, child_qty, db, visited).items():
                result[mat_id] = result.get(mat_id, 0) + amt

    return result


def explode_bom(product_id: int, produced_qty: float, db: Session,
                visited: frozenset = frozenset()) -> dict:
    """
    완제품 BOM을 재귀 전개해 원재료별 차감량 반환.
    - material_id: 직접 차감
    - semi_product_id: 반제품 BOM 재귀 전개 (explode_semi_bom)
    - child_product_id: 완제품 BOM 재귀 전개 (explode_bom)
    반환: {material_id: deduct_amount}
    """
    if product_id in visited:
        return {}
    visited = visited | {product_id}

    product = db.query(FinishedProduct).get(product_id)
    if not product:
        return {}

    eff = _conv_eff(product, produced_qty)
    result: dict = {}

    for bom in db.query(ProductBOM).filter(ProductBOM.product_id == product_id).all():
        if bom.material_id:
            amt = eff * (bom.quantity / 100) if bom.unit == '%' else bom.quantity * produced_qty
            result[bom.material_id] = result.get(bom.material_id, 0) + amt

        elif bom.semi_product_id:
            # 반제품 BOM 재귀 전개
            semi_qty = eff * (bom.quantity / 100) if bom.unit == '%' else bom.quantity * produced_qty
            for mat_id, amt in explode_semi_bom(bom.semi_product_id, semi_qty, db).items():
                result[mat_id] = result.get(mat_id, 0) + amt

        elif bom.child_product_id:
            # 완제품 BOM 재귀 전개
            child_qty = eff * (bom.quantity / 100) if bom.unit == '%' else bom.quantity * produced_qty
            for mat_id, amt in explode_bom(bom.child_product_id, child_qty, db, visited).items():
                result[mat_id] = result.get(mat_id, 0) + amt

    return result


# ─────────────────────────────────────────
# Materials
# ─────────────────────────────────────────
@app.get("/api/materials")
def list_materials(db: Session = Depends(get_db)):
    items = db.query(Material).filter(Material.user_id == uid()).all()
    result = []
    for m in items:
        status = "normal"
        if m.current_stock <= 0:
            status = "danger"
        elif m.current_stock < m.safety_stock:
            status = "danger"
        elif m.current_stock < m.safety_stock * 1.5:
            status = "warn"
        result.append({
            "id": m.id, "name": m.name, "unit": m.unit,
            "safety_stock": m.safety_stock, "current_stock": m.current_stock,
            "status": status,
        })
    return result

@app.post("/api/materials", status_code=201)
def create_material(data: MaterialCreate, db: Session = Depends(get_db)):
    m = Material(**data.model_dump(), user_id=uid())
    db.add(m); db.commit(); db.refresh(m)
    return {"id": m.id, "name": m.name}

@app.patch("/api/materials/{material_id}")
def update_material(material_id: int, data: MaterialUpdate, db: Session = Depends(get_db)):
    m = db.query(Material).filter(Material.id == material_id, Material.user_id == uid()).first()
    if not m:
        raise HTTPException(404, "Not found")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(m, k, v)
    db.commit()
    return {"ok": True}


# ─────────────────────────────────────────
# Suppliers
# ─────────────────────────────────────────
@app.get("/api/suppliers")
def list_suppliers(db: Session = Depends(get_db)):
    items = db.query(Supplier).filter(Supplier.user_id == uid()).all()
    result = []
    for s in items:
        this_month_receipts = [
            r for r in s.receipts
            if r.created_at and r.created_at.month == datetime.now().month
        ]
        result.append({
            "id": s.id, "name": s.name, "business_number": s.business_number,
            "status": s.status, "ocr_mapped": s.ocr_mapped,
            "monthly_receipts": len(this_month_receipts),
        })
    return result

@app.post("/api/suppliers", status_code=201)
def create_supplier(data: SupplierCreate, db: Session = Depends(get_db)):
    s = Supplier(**data.model_dump(), user_id=uid())
    db.add(s); db.commit(); db.refresh(s)
    return {"id": s.id, "name": s.name}


# ─────────────────────────────────────────
# Recipes
# ─────────────────────────────────────────
@app.get("/api/recipes")
def list_recipes(db: Session = Depends(get_db)):
    items = db.query(Recipe).filter(Recipe.user_id == uid()).all()
    result = []
    for r in items:
        ingredients = []
        for ing in r.ingredients:
            ingredients.append({
                "material_id": ing.material_id,
                "material_name": ing.material.name if ing.material else "",
                "quantity": ing.quantity,
                "unit": ing.unit,
            })
        result.append({
            "id": r.id, "product_code": r.product_code,
            "product_name": r.product_name, "category": r.category,
            "base_quantity": r.base_quantity, "base_unit": r.base_unit,
            "version": r.version, "status": r.status,
            "ingredients": ingredients,
        })
    return result

@app.post("/api/recipes", status_code=201)
def create_recipe(data: RecipeCreate, db: Session = Depends(get_db)):
    r = Recipe(**data.model_dump(), user_id=uid())
    db.add(r); db.commit(); db.refresh(r)
    return {"id": r.id, "product_code": r.product_code}

@app.post("/api/recipes/{recipe_id}/ingredients", status_code=201)
def add_ingredient(recipe_id: int, data: RecipeIngredientCreate, db: Session = Depends(get_db)):
    r = db.query(Recipe).get(recipe_id)
    if not r:
        raise HTTPException(404, "Recipe not found")
    ing = RecipeIngredient(recipe_id=recipe_id, **data.model_dump())
    db.add(ing); db.commit()
    return {"ok": True}


# ─────────────────────────────────────────
# Receipts (원료 입고)
# ─────────────────────────────────────────
@app.get("/api/receipts")
def list_receipts(
    limit: int = 50,
    material_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    q = db.query(Receipt).filter(Receipt.user_id == uid()).order_by(desc(Receipt.created_at))
    if material_id:
        q = q.filter(Receipt.material_id == material_id)
    items = q.limit(limit).all()
    result = []
    for r in items:
        result.append({
            "id": r.id, "receipt_number": r.receipt_number,
            "material_name": r.material.name if r.material else "",
            "supplier_name": r.supplier.name if r.supplier else "",
            "lot_number": r.lot_number, "quantity": r.quantity,
            "unit": r.material.unit if r.material else "",
            "unit_price": r.unit_price,
            "delivery_date": r.delivery_date.strftime("%Y-%m-%d") if r.delivery_date else None,
            "expiry_date": r.expiry_date.strftime("%Y-%m-%d") if r.expiry_date else None,
            "input_method": r.input_method, "status": r.status,
            "packaging_ok": r.packaging_ok, "visual_ok": r.visual_ok,
            "judgment_ok": r.judgment_ok, "inspector": r.inspector, "confirmer": r.confirmer,
        })
    return result

@app.post("/api/receipts", status_code=201)
def create_receipt(data: ReceiptCreate, db: Session = Depends(get_db)):
    material = db.query(Material).filter(Material.id == data.material_id, Material.user_id == uid()).first()
    if not material:
        raise HTTPException(404, "Material not found")

    now = datetime.now()
    receipt_number = f"R-{now.strftime('%Y%m%d')}-{secrets.token_hex(2).upper()}"

    receipt = Receipt(
        receipt_number=receipt_number,
        material_id=data.material_id,
        supplier_id=data.supplier_id,
        lot_number=data.lot_number or f"LOT-{now.strftime('%Y%m%d')}-AUTO",
        quantity=data.quantity,
        unit_price=data.unit_price,
        delivery_date=datetime.fromisoformat(data.delivery_date) if data.delivery_date else now,
        expiry_date=datetime.fromisoformat(data.expiry_date) if data.expiry_date else None,
        input_method=data.input_method,
        status="confirmed",
        user_id=uid(),
        packaging_ok=data.packaging_ok,
        visual_ok=data.visual_ok,
        judgment_ok=data.judgment_ok,
        inspector=data.inspector,
        confirmer=data.confirmer,
    )
    db.add(receipt)
    material.current_stock += data.quantity
    db.commit()
    db.refresh(receipt)
    return {"id": receipt.id, "receipt_number": receipt.receipt_number}


@app.get("/api/receipts/ingodaejang")
def get_ingodaejang(
    material_id: int,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db)
):
    material = db.query(Material).filter(Material.id == material_id, Material.user_id == uid()).first()
    if not material:
        raise HTTPException(404, "Material not found")
    q = db.query(Receipt).filter(Receipt.material_id == material_id, Receipt.user_id == uid())
    if date_from:
        q = q.filter(Receipt.delivery_date >= datetime.fromisoformat(date_from))
    if date_to:
        q = q.filter(Receipt.delivery_date <= datetime.fromisoformat(date_to + "T23:59:59"))
    rows = q.order_by(Receipt.delivery_date).all()
    nc_all = db.query(ReceiptNonconformance).filter(
        ReceiptNonconformance.user_id == uid(),
        ReceiptNonconformance.receipt_id.in_([r.id for r in rows])
    ).order_by(ReceiptNonconformance.date).all()
    return {
        "material_name": material.name,
        "rows": [{
            "id": r.id,
            "delivery_date": r.delivery_date.strftime("%Y-%m-%d") if r.delivery_date else "",
            "supplier_name": r.supplier.name if r.supplier else "",
            "packaging_ok": r.packaging_ok,
            "expiry_date": r.expiry_date.strftime("%Y-%m-%d") if r.expiry_date else "",
            "visual_ok": r.visual_ok,
            "judgment_ok": r.judgment_ok,
            "inspector": r.inspector or "",
            "confirmer": r.confirmer or "",
        } for r in rows],
        "nonconformances": [{
            "date": nc.date or "",
            "content": nc.content or "",
            "inspector": nc.inspector or "",
        } for nc in nc_all],
    }

@app.post("/api/receipts/nonconformance", status_code=201)
def add_nonconformance(data: NonconformanceCreate, db: Session = Depends(get_db)):
    r = db.query(Receipt).filter(Receipt.id == data.receipt_id, Receipt.user_id == uid()).first()
    if not r:
        raise HTTPException(404, "Receipt not found")
    nc = ReceiptNonconformance(
        receipt_id=data.receipt_id,
        date=data.date,
        content=data.content,
        inspector=data.inspector,
        user_id=uid(),
    )
    db.add(nc); db.commit(); db.refresh(nc)
    return {"id": nc.id}

@app.delete("/api/receipts/{receipt_id}", status_code=200)
def delete_receipt(receipt_id: int, db: Session = Depends(get_db)):
    r = db.query(Receipt).filter(Receipt.id == receipt_id, Receipt.user_id == uid()).first()
    if not r:
        raise HTTPException(404, "Not found")
    mat = db.query(Material).filter(Material.id == r.material_id, Material.user_id == uid()).first()
    if mat:
        mat.current_stock = max(0, mat.current_stock - r.quantity)
    db.delete(r)
    db.commit()
    return {"ok": True}


# ─────────────────────────────────────────
# Ledger (수불대장)
# ─────────────────────────────────────────
@app.get("/api/ledger")
def get_ledger(
    material_id: int,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db)
):
    material = db.query(Material).filter(Material.id == material_id, Material.user_id == uid()).first()
    if not material:
        raise HTTPException(404, "Material not found")

    receipts_q = db.query(Receipt).filter(Receipt.user_id == uid(), Receipt.material_id == material_id, Receipt.status == "confirmed")
    if date_from:
        receipts_q = receipts_q.filter(Receipt.delivery_date >= datetime.fromisoformat(date_from))
    if date_to:
        receipts_q = receipts_q.filter(Receipt.delivery_date <= datetime.fromisoformat(date_to))

    entries = []
    for r in receipts_q.order_by(Receipt.delivery_date).all():
        entries.append({
            "date": r.delivery_date.strftime("%Y-%m-%d") if r.delivery_date else "",
            "type": "입고",
            "in_qty": r.quantity, "out_qty": 0,
            "supplier": r.supplier.name if r.supplier else "",
            "lot": r.lot_number or "",
            "input_method": r.input_method,
        })

    # BOM 기반 생산차감 조회
    bom_prods_q = db.query(Production).join(
        ProductBOM, ProductBOM.product_id == Production.finished_product_id
    ).filter(Production.user_id == uid(), ProductBOM.material_id == material_id)
    if date_from:
        bom_prods_q = bom_prods_q.filter(Production.start_time >= datetime.fromisoformat(date_from))
    if date_to:
        bom_prods_q = bom_prods_q.filter(Production.start_time <= datetime.fromisoformat(date_to))

    def to_base_kg_led(qty, unit):
        if unit in ('g', 'ml'): return qty / 1000
        return qty

    for p in bom_prods_q.order_by(Production.start_time).all():
        prod_qty = p.produced_quantity or 0
        fp = p.finished_product
        # 단위환산
        if fp and fp.unit_conv_qty and fp.unit_conv_unit:
            u1 = fp.unit_conv_unit
            if u1 in ('kg', 'g', 'L', 'ml'):
                eff_kg = prod_qty * to_base_kg_led(fp.unit_conv_qty, u1)
            elif u1 == 'ea' and fp.unit_conv2_qty and fp.unit_conv2_unit:
                eff_kg = prod_qty * fp.unit_conv_qty * to_base_kg_led(fp.unit_conv2_qty, fp.unit_conv2_unit)
            else:
                eff_kg = prod_qty
        else:
            eff_kg = prod_qty

        bom = db.query(ProductBOM).filter(
            ProductBOM.product_id == p.finished_product_id,
            ProductBOM.material_id == material_id
        ).first()
        if bom:
            if bom.unit == '%':
                usage = eff_kg * (bom.quantity / 100)
            else:
                usage = bom.quantity * prod_qty
            entries.append({
                "date": p.start_time.strftime("%Y-%m-%d") if p.start_time else "",
                "type": "생산사용",
                "in_qty": 0, "out_qty": round(usage, 3),
                "supplier": fp.name if fp else "",
                "lot": p.lot_number or "",
                "input_method": p.input_method,
            })

    # 레거시 레시피 기반 생산차감
    recipe_prods_q = db.query(Production).join(Recipe).join(RecipeIngredient).filter(
        Production.user_id == uid(),
        RecipeIngredient.material_id == material_id,
        Production.finished_product_id == None
    )
    if date_from:
        recipe_prods_q = recipe_prods_q.filter(Production.start_time >= datetime.fromisoformat(date_from))
    if date_to:
        recipe_prods_q = recipe_prods_q.filter(Production.start_time <= datetime.fromisoformat(date_to))

    for p in recipe_prods_q.order_by(Production.start_time).all():
        usage = sum(
            ing.quantity * (p.produced_quantity / p.recipe.base_quantity)
            for ing in p.recipe.ingredients
            if ing.material_id == material_id
        ) if p.produced_quantity and p.recipe and p.recipe.base_quantity else 0
        entries.append({
            "date": p.start_time.strftime("%Y-%m-%d") if p.start_time else "",
            "type": "생산사용",
            "in_qty": 0, "out_qty": round(usage, 3),
            "supplier": p.lot_number or "",
            "lot": "",
            "input_method": p.input_method,
        })

    entries.sort(key=lambda x: x["date"])

    running = 0
    for e in entries:
        running += e["in_qty"] - e["out_qty"]
        e["balance"] = round(running, 3)

    total_in = sum(e["in_qty"] for e in entries)
    total_out = sum(e["out_qty"] for e in entries)

    return {
        "material_name": material.name,
        "unit": material.unit,
        "total_in": round(total_in, 3),
        "total_out": round(total_out, 3),
        "current_stock": round(material.current_stock, 3),
        "entries": entries,
    }


# ─────────────────────────────────────────
# 수불대장 전체 (날짜 기준)
# ─────────────────────────────────────────
@app.get("/api/ledger/all")
def get_ledger_all(
    date_from: str,
    date_to: str,
    db: Session = Depends(get_db)
):
    dt_from = datetime.fromisoformat(date_from)
    dt_to = datetime.fromisoformat(date_to) + timedelta(days=1)

    _uid = uid()
    materials = db.query(Material).filter(Material.user_id == _uid, Material.status == 'active').order_by(Material.name).all()
    mat_map = {m.id: m for m in materials}

    def to_base(qty, unit):
        if unit in ('g', 'ml'): return qty / 1000
        return qty

    def compute_opening(mid):
        """기간 시작일 이전 누적 잔고 (재귀 BOM 전개)"""
        recv = db.query(func.sum(Receipt.quantity)).filter(
            Receipt.user_id == _uid,
            Receipt.material_id == mid,
            Receipt.status == 'confirmed',
            Receipt.delivery_date < dt_from
        ).scalar() or 0
        adj = db.query(func.sum(StockAdjustment.diff_qty)).filter(
            StockAdjustment.user_id == _uid,
            StockAdjustment.material_id == mid,
            StockAdjustment.adjusted_at < dt_from
        ).scalar() or 0
        usage = 0.0
        prods = db.query(Production).filter(
            Production.user_id == _uid,
            Production.finished_product_id.isnot(None),
            Production.start_time < dt_from
        ).all()
        seen = set()
        for p in prods:
            if p.id in seen: continue
            seen.add(p.id)
            pq = p.produced_quantity or 0
            if pq <= 0: continue
            deductions = explode_bom(p.finished_product_id, pq, db)
            usage += deductions.get(mid, 0)
        return recv - usage + adj

    opening = {m.id: compute_opening(m.id) for m in materials}
    all_entries = []

    # 입고
    for r in db.query(Receipt).filter(
        Receipt.user_id == _uid,
        Receipt.status == 'confirmed',
        Receipt.delivery_date >= dt_from,
        Receipt.delivery_date < dt_to
    ).all():
        if r.material_id not in mat_map: continue
        m = mat_map[r.material_id]
        all_entries.append({
            'dt': r.delivery_date or datetime.min,
            'date': (r.delivery_date or datetime.min).strftime('%Y-%m-%d'),
            'mid': m.id, 'mat': m.name, 'unit': m.unit,
            'type': '입고', 'source': r.supplier.name if r.supplier else '',
            'in_qty': r.quantity, 'out_qty': 0,
            'note': r.lot_number or '',
        })

    # 생산차감 (재귀 BOM 전개)
    seen_prod_mat = set()
    for p in db.query(Production).filter(
        Production.user_id == _uid,
        Production.finished_product_id.isnot(None),
        Production.start_time >= dt_from,
        Production.start_time < dt_to
    ).all():
        pq = p.produced_quantity or 0
        if pq <= 0: continue
        fp = p.finished_product
        deductions = explode_bom(p.finished_product_id, pq, db)
        for mat_id, usage in deductions.items():
            if mat_id not in mat_map: continue
            key = (p.id, mat_id)
            if key in seen_prod_mat: continue
            seen_prod_mat.add(key)
            m = mat_map[mat_id]
            all_entries.append({
                'dt': p.start_time or datetime.min,
                'date': (p.start_time or datetime.min).strftime('%Y-%m-%d'),
                'mid': m.id, 'mat': m.name, 'unit': m.unit,
                'type': '생산사용', 'source': fp.name if fp else '',
                'in_qty': 0, 'out_qty': round(usage, 3),
                'note': p.lot_number or '',
            })

    # 재고조정
    for a in db.query(StockAdjustment).filter(
        StockAdjustment.user_id == _uid,
        StockAdjustment.adjusted_at >= dt_from,
        StockAdjustment.adjusted_at < dt_to
    ).all():
        if a.material_id not in mat_map: continue
        m = mat_map[a.material_id]
        all_entries.append({
            'dt': a.adjusted_at,
            'date': a.adjusted_at.strftime('%Y-%m-%d'),
            'mid': m.id, 'mat': m.name, 'unit': m.unit,
            'type': '재고조정', 'source': a.reason or '',
            'in_qty': a.diff_qty if a.diff_qty > 0 else 0,
            'out_qty': abs(a.diff_qty) if a.diff_qty < 0 else 0,
            'note': a.note or '',
        })

    all_entries.sort(key=lambda x: (x['date'], x['mid']))

    balances = dict(opening)
    result = []
    for e in all_entries:
        balances[e['mid']] = round(balances[e['mid']] + e['in_qty'] - e['out_qty'], 3)
        result.append({
            'date': e['date'], 'material_name': e['mat'], 'unit': e['unit'],
            'type': e['type'], 'source': e['source'],
            'in_qty': e['in_qty'], 'out_qty': e['out_qty'],
            'balance': balances[e['mid']], 'note': e['note'],
        })

    return {
        'date_from': date_from, 'date_to': date_to,
        'total_in': round(sum(e['in_qty'] for e in all_entries), 3),
        'total_out': round(sum(e['out_qty'] for e in all_entries), 3),
        'entries': result,
    }


# ─────────────────────────────────────────
# 재고조정
# ─────────────────────────────────────────
class StockAdjustmentIn(BaseModel):
    material_id: int
    actual_qty: float
    reason: str
    note: Optional[str] = None

@app.post("/api/stock-adjustments", status_code=201)
def create_stock_adjustment(data: StockAdjustmentIn, db: Session = Depends(get_db)):
    m = db.query(Material).filter(Material.id == data.material_id, Material.user_id == uid()).first()
    if not m:
        raise HTTPException(404, "Not found")
    before = m.current_stock
    diff = round(data.actual_qty - before, 3)
    adj = StockAdjustment(
        material_id=data.material_id,
        before_qty=before,
        after_qty=data.actual_qty,
        diff_qty=diff,
        reason=data.reason,
        note=data.note,
        user_id=uid(),
    )
    m.current_stock = data.actual_qty
    db.add(adj)
    db.commit()
    return {"id": adj.id, "before": before, "after": data.actual_qty, "diff": diff}

@app.get("/api/stock-adjustments")
def list_stock_adjustments(material_id: Optional[int] = None, db: Session = Depends(get_db)):
    q = db.query(StockAdjustment).filter(StockAdjustment.user_id == uid())
    if material_id:
        q = q.filter(StockAdjustment.material_id == material_id)
    return [{
        "id": a.id,
        "material_name": a.material.name if a.material else "",
        "unit": a.material.unit if a.material else "",
        "before_qty": a.before_qty,
        "after_qty": a.after_qty,
        "diff_qty": a.diff_qty,
        "reason": a.reason,
        "note": a.note,
        "adjusted_at": a.adjusted_at.strftime("%Y-%m-%d %H:%M") if a.adjusted_at else "",
    } for a in q.order_by(StockAdjustment.adjusted_at.desc()).all()]


# ─────────────────────────────────────────
# Productions (생산 기록)
# ─────────────────────────────────────────
@app.get("/api/productions")
def list_productions(
    limit: int = 50,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    q = db.query(Production).filter(Production.user_id == uid()).order_by(desc(Production.created_at))
    if date_from:
        q = q.filter(Production.start_time >= datetime.fromisoformat(date_from))
    if date_to:
        q = q.filter(Production.start_time <= datetime.fromisoformat(date_to))
    if status:
        q = q.filter(Production.status == status)
    items = q.limit(limit).all()
    result = []
    for p in items:
        good_rate = round(p.good_quantity / p.produced_quantity * 100, 1) if p.produced_quantity and p.good_quantity else None
        # finished_product 우선, fallback to recipe
        if p.finished_product_id and p.finished_product:
            product_name = p.finished_product.name
            product_unit = p.finished_product.unit
        elif p.recipe:
            product_name = p.recipe.product_name
            product_unit = p.recipe.base_unit
        else:
            product_name = ""
            product_unit = ""
        result.append({
            "id": p.id, "lot_number": p.lot_number,
            "product_name": product_name, "product_unit": product_unit,
            "finished_product_id": p.finished_product_id,
            "device_name": p.device.name if p.device else "",
            "produced_quantity": p.produced_quantity,
            "good_quantity": p.good_quantity,
            "defect_quantity": p.defect_quantity,
            "good_rate": good_rate,
            "start_time": p.start_time.strftime("%Y-%m-%d %H:%M") if p.start_time else None,
            "end_time": p.end_time.strftime("%Y-%m-%d %H:%M") if p.end_time else None,
            "input_method": p.input_method,
            "status": p.status,
        })
    return result

@app.get("/api/productions/bom-preview")
def production_bom_preview(product_id: int, qty: float, db: Session = Depends(get_db)):
    """BOM 재귀 전개 후 차감 예정 원재료 목록 반환"""
    deductions = explode_bom(product_id, qty, db)
    result = []
    for mat_id, amt in deductions.items():
        mat = db.query(Material).get(mat_id)
        if mat:
            result.append({"material_id": mat_id, "material_name": mat.name,
                           "unit": mat.unit, "deduct_qty": round(amt, 3)})
    return result

@app.get("/api/debug/semi-bom")
def debug_semi_bom(semi_id: int, qty: float, db: Session = Depends(get_db)):
    semi = db.query(SemiProduct).get(semi_id)
    rows = db.query(SemiProductBOM).filter(SemiProductBOM.semi_product_id == semi_id).all()
    deductions = explode_semi_bom(semi_id, qty, db)
    return {
        "semi_found": semi is not None,
        "semi_name": semi.name if semi else None,
        "bom_rows": len(rows),
        "bom_detail": [{"material_id": b.material_id, "qty": b.quantity, "unit": b.unit} for b in rows],
        "deductions": deductions,
    }


@app.get("/api/productions/{prod_id}")
def get_production(prod_id: int, db: Session = Depends(get_db)):
    p = db.query(Production).get(prod_id)
    if not p:
        raise HTTPException(404, "Not found")
    if p.finished_product_id and p.finished_product:
        product_name, product_unit = p.finished_product.name, p.finished_product.unit
    elif p.recipe:
        product_name, product_unit = p.recipe.product_name, p.recipe.base_unit
    else:
        product_name = product_unit = ""
    return {
        "id": p.id, "lot_number": p.lot_number,
        "finished_product_id": p.finished_product_id,
        "product_name": product_name, "product_unit": product_unit,
        "produced_quantity": p.produced_quantity,
        "good_quantity": p.good_quantity,
        "defect_quantity": p.defect_quantity,
        "start_time": p.start_time.strftime("%Y-%m-%dT%H:%M") if p.start_time else None,
        "end_time": p.end_time.strftime("%Y-%m-%dT%H:%M") if p.end_time else None,
        "note": p.note, "status": p.status, "input_method": p.input_method,
    }


class ProductionUpdate(BaseModel):
    finished_product_id: Optional[int] = None
    produced_quantity: Optional[float] = None
    defect_quantity: Optional[float] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    note: Optional[str] = None


@app.patch("/api/productions/{prod_id}")
def update_production(prod_id: int, data: ProductionUpdate, db: Session = Depends(get_db)):
    prod = db.query(Production).filter(Production.id == prod_id, Production.user_id == uid()).first()
    if not prod:
        raise HTTPException(404, "Not found")

    # 기존 차감 되돌리기
    if prod.finished_product_id and prod.produced_quantity:
        for mat_id, amt in explode_bom(prod.finished_product_id, prod.produced_quantity, db).items():
            mat = db.query(Material).filter(Material.id == mat_id, Material.user_id == uid()).first()
            if mat:
                mat.current_stock = mat.current_stock + round(amt, 3)
        # 완제품 재고 역산
        old_good = (prod.good_quantity or prod.produced_quantity or 0)
        fp_old = db.query(FinishedProduct).filter(FinishedProduct.id == prod.finished_product_id, FinishedProduct.user_id == uid()).first()
        if fp_old:
            fp_old.current_stock = max(0, (fp_old.current_stock or 0) - old_good)

    new_pid = data.finished_product_id if data.finished_product_id is not None else prod.finished_product_id
    new_qty = data.produced_quantity if data.produced_quantity is not None else prod.produced_quantity
    new_defect = data.defect_quantity if data.defect_quantity is not None else (prod.defect_quantity or 0)

    if data.finished_product_id is not None: prod.finished_product_id = data.finished_product_id
    if data.produced_quantity is not None:   prod.produced_quantity   = data.produced_quantity
    if data.defect_quantity is not None:     prod.defect_quantity     = data.defect_quantity
    if data.start_time is not None:
        prod.start_time = datetime.fromisoformat(data.start_time) if data.start_time else None
    if data.end_time is not None:
        prod.end_time = datetime.fromisoformat(data.end_time) if data.end_time else None
    if data.note is not None: prod.note = data.note

    prod.good_quantity = (new_qty or 0) - new_defect

    # 새 차감 적용
    if new_pid and new_qty:
        for mat_id, amt in explode_bom(new_pid, new_qty, db).items():
            mat = db.query(Material).filter(Material.id == mat_id, Material.user_id == uid()).first()
            if mat:
                mat.current_stock = max(0, mat.current_stock - round(amt, 3))
        new_good = new_qty - new_defect
        fp_new = db.query(FinishedProduct).filter(FinishedProduct.id == new_pid, FinishedProduct.user_id == uid()).first()
        if fp_new:
            fp_new.current_stock = (fp_new.current_stock or 0) + new_good

    db.commit()
    return {"id": prod.id}


@app.post("/api/productions", status_code=201)
def create_production(data: ProductionCreate, db: Session = Depends(get_db)):
    now = datetime.now()
    lot = data.lot_number or f"PROD-{now.strftime('%Y%m%d')}-{secrets.token_hex(2).upper()}"

    defect = data.defect_quantity or 0
    good = data.good_quantity if data.good_quantity is not None else (
        (data.produced_quantity or 0) - defect
    )
    prod = Production(
        lot_number=lot,
        finished_product_id=data.finished_product_id,
        recipe_id=data.recipe_id,
        device_id=data.device_id,
        produced_quantity=data.produced_quantity,
        good_quantity=good,
        defect_quantity=defect,
        start_time=datetime.fromisoformat(data.start_time) if data.start_time else now,
        end_time=datetime.fromisoformat(data.end_time) if data.end_time else now,
        input_method=data.input_method,
        status=data.status,
        note=data.note,
        user_id=uid(),
    )
    db.add(prod)

    # BOM 기반 원재료 차감 (재귀 전개)
    if data.finished_product_id and data.produced_quantity:
        deductions = explode_bom(data.finished_product_id, data.produced_quantity, db)
        for mat_id, amt in deductions.items():
            mat = db.query(Material).filter(Material.id == mat_id, Material.user_id == uid()).first()
            if mat:
                mat.current_stock = max(0, mat.current_stock - round(amt, 3))
        # 완제품 재고 증가
        fp = db.query(FinishedProduct).filter(FinishedProduct.id == data.finished_product_id, FinishedProduct.user_id == uid()).first()
        if fp:
            fp.current_stock = (fp.current_stock or 0) + (good)
    # 레거시: recipe 기반 차감
    elif data.recipe_id and data.produced_quantity:
        recipe = db.query(Recipe).filter(Recipe.id == data.recipe_id, Recipe.user_id == uid()).first()
        if recipe and recipe.base_quantity:
            ratio = data.produced_quantity / recipe.base_quantity
            for ing in recipe.ingredients:
                mat = db.query(Material).filter(Material.id == ing.material_id, Material.user_id == uid()).first()
                if mat:
                    mat.current_stock = max(0, mat.current_stock - ing.quantity * ratio)

    db.commit()
    db.refresh(prod)
    return {"id": prod.id, "lot_number": prod.lot_number}


# ─────────────────────────────────────────
# AI 추천: 생산실적 누락 시 과거 데이터 기반 추천
# ─────────────────────────────────────────
@app.get("/api/productions/recommend")
def recommend_production(
    recipe_id: int,
    date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    과거 동일 레시피의 생산 실적을 분석해 누락된 생산량/양품률을 추천합니다.
    - 최근 30일 데이터 기반
    - 요일(day_of_week) 패턴 반영
    - 이상치 제거 후 중앙값 사용
    """
    recipe = db.query(Recipe).filter(Recipe.id == recipe_id, Recipe.user_id == uid()).first()
    if not recipe:
        raise HTTPException(404, "Recipe not found")

    target_date = datetime.fromisoformat(date) if date else datetime.now()
    target_dow = target_date.weekday()  # 0=월 ~ 6=일

    cutoff = target_date - timedelta(days=30)
    past = db.query(Production).filter(
        Production.user_id == uid(),
        Production.recipe_id == recipe_id,
        Production.status == "completed",
        Production.produced_quantity != None,
        Production.start_time >= cutoff,
    ).order_by(desc(Production.start_time)).all()

    if not past:
        return {
            "recipe_id": recipe_id,
            "recipe_name": recipe.product_name,
            "recommended": False,
            "message": "과거 데이터가 없어 추천할 수 없습니다.",
        }

    # 요일 일치 데이터
    same_dow = [p for p in past if p.start_time and p.start_time.weekday() == target_dow]
    base_data = same_dow if len(same_dow) >= 3 else past

    quantities = [p.produced_quantity for p in base_data if p.produced_quantity]
    good_rates = [
        p.good_quantity / p.produced_quantity * 100
        for p in base_data
        if p.produced_quantity and p.good_quantity
    ]

    # IQR 이상치 제거
    def iqr_filter(values):
        if len(values) < 4:
            return values
        values_sorted = sorted(values)
        q1 = values_sorted[len(values_sorted) // 4]
        q3 = values_sorted[3 * len(values_sorted) // 4]
        iqr = q3 - q1
        return [v for v in values if q1 - 1.5 * iqr <= v <= q3 + 1.5 * iqr]

    clean_qty = iqr_filter(quantities)
    clean_rate = iqr_filter(good_rates)

    rec_qty = round(statistics.median(clean_qty)) if clean_qty else None
    rec_rate = round(statistics.mean(clean_rate), 1) if clean_rate else None
    rec_good = round(rec_qty * rec_rate / 100) if rec_qty and rec_rate else None
    rec_defect = (rec_qty - rec_good) if rec_qty and rec_good else None

    # 최근 7일 추이
    recent_7 = past[:7]
    trend = "stable"
    if len(recent_7) >= 3:
        half = len(recent_7) // 2
        early_avg = statistics.mean([p.produced_quantity for p in recent_7[half:] if p.produced_quantity] or [0])
        late_avg  = statistics.mean([p.produced_quantity for p in recent_7[:half]  if p.produced_quantity] or [0])
        if late_avg > early_avg * 1.05:
            trend = "up"
        elif late_avg < early_avg * 0.95:
            trend = "down"

    return {
        "recipe_id": recipe_id,
        "recipe_name": recipe.product_name,
        "target_date": target_date.strftime("%Y-%m-%d"),
        "day_of_week": ["월", "화", "수", "목", "금", "토", "일"][target_dow],
        "recommended": True,
        "recommended_produced_quantity": rec_qty,
        "recommended_good_quantity": rec_good,
        "recommended_defect_quantity": rec_defect,
        "recommended_good_rate": rec_rate,
        "based_on_records": len(base_data),
        "same_dow_records": len(same_dow),
        "trend": trend,
        "message": (
            f"최근 {len(base_data)}건의 {recipe.product_name} 생산 실적"
            + (f" (동일 요일 {len(same_dow)}건 포함)" if same_dow else "")
            + f" 기반 추천값입니다."
        ),
    }


# ─────────────────────────────────────────
# OCR — 거래명세서 자동 분석
# ─────────────────────────────────────────
def _normalize(text: str) -> str:
    return re.sub(r'[\s\(\)\(\)주식회사(주)]', '', text).lower()

@app.post("/api/ocr/receipt")
async def ocr_receipt(image: UploadFile = File(...), db: Session = Depends(get_db)):
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key or api_key == "여기에_API_키를_입력하세요":
        raise HTTPException(500, "C:\\foodly\\.env 파일을 열어 ANTHROPIC_API_KEY에 실제 키를 입력한 뒤 서버를 재시작하세요.")

    img_data = await image.read()
    img_b64 = base64.standard_b64encode(img_data).decode("utf-8")
    media_type = image.content_type or "image/jpeg"
    if media_type not in ("image/jpeg", "image/png", "image/gif", "image/webp"):
        media_type = "image/jpeg"

    try:
        try:
            import anthropic as _anthropic
        except ModuleNotFoundError:
            import subprocess, sys
            subprocess.check_call([sys.executable, "-m", "pip", "install", "anthropic", "-q"])
            import anthropic as _anthropic
        client = _anthropic.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1500,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": img_b64}},
                    {"type": "text", "text": (
                        "이 거래명세서 이미지에서 정보를 추출해주세요.\n"
                        "반드시 아래 JSON 형식으로만 응답하세요 (마크다운/설명 없이):\n"
                        '{"supplier_name":"공급업체명 또는 null","delivery_date":"YYYY-MM-DD 또는 null",'
                        '"items":[{"name":"품목명","quantity":수량숫자,"unit":"kg/g/L/ml/ea 등",'
                        '"unit_price":단가숫자또는null,"expiry_date":"YYYY-MM-DD 또는 null","lot_number":"로트번호 또는 null"}]}'
                    )},
                ],
            }],
        )
        raw = msg.content[0].text.strip()
        # strip markdown code fences if present
        raw = re.sub(r'^```[a-z]*\s*', '', raw); raw = re.sub(r'\s*```$', '', raw)
        extracted = json.loads(raw)
    except Exception as e:
        raise HTTPException(500, f"OCR 분석 실패: {str(e)}")

    suppliers = db.query(Supplier).filter(Supplier.user_id == uid(), Supplier.status == "active").all()
    materials = db.query(Material).filter(Material.user_id == uid(), Material.status == "active").all()

    # match supplier
    supplier_name = (extracted.get("supplier_name") or "").strip()
    matched_supplier = None
    if supplier_name:
        q = _normalize(supplier_name)
        for s in suppliers:
            if q in _normalize(s.name) or _normalize(s.name) in q:
                matched_supplier = s; break

    # match materials
    matched_items = []
    for item in extracted.get("items", []):
        item_name = (item.get("name") or "").strip()
        matched_mat = None
        if item_name:
            q = _normalize(item_name)
            for m in materials:
                if q in _normalize(m.name) or _normalize(m.name) in q:
                    matched_mat = m; break
        matched_items.append({
            "extracted_name": item_name,
            "matched_material_id": matched_mat.id if matched_mat else None,
            "matched_material_name": matched_mat.name if matched_mat else None,
            "matched_unit": matched_mat.unit if matched_mat else item.get("unit") or "kg",
            "quantity": item.get("quantity"),
            "unit": item.get("unit") or (matched_mat.unit if matched_mat else "kg"),
            "unit_price": item.get("unit_price"),
            "expiry_date": item.get("expiry_date"),
            "lot_number": item.get("lot_number"),
        })

    return {
        "supplier_name": supplier_name,
        "matched_supplier_id": matched_supplier.id if matched_supplier else None,
        "matched_supplier_name": matched_supplier.name if matched_supplier else None,
        "delivery_date": extracted.get("delivery_date"),
        "items": matched_items,
    }


# ─────────────────────────────────────────
# Dashboard stats
# ─────────────────────────────────────────
@app.get("/api/dashboard")
def get_dashboard(db: Session = Depends(get_db)):
    today = datetime.now().date()
    yesterday = today - timedelta(days=1)

    _uid = uid()
    today_prods = db.query(Production).filter(
        Production.user_id == _uid,
        func.date(Production.start_time) == today
    ).all()
    yesterday_prods = db.query(Production).filter(
        Production.user_id == _uid,
        func.date(Production.start_time) == yesterday
    ).all()

    today_qty = sum(p.produced_quantity or 0 for p in today_prods)
    today_good = sum(p.good_quantity or 0 for p in today_prods)
    today_rate = round(today_good / today_qty * 100, 1) if today_qty else 0

    yesterday_qty = sum(p.produced_quantity or 0 for p in yesterday_prods)
    yesterday_good = sum(p.good_quantity or 0 for p in yesterday_prods)
    yesterday_rate = round(yesterday_good / yesterday_qty * 100, 1) if yesterday_qty else 0

    today_receipt_rows = db.query(Receipt).filter(
        Receipt.user_id == _uid,
        func.date(Receipt.delivery_date) == today
    ).all()
    today_receipts = len(today_receipt_rows)
    today_receipt_qty = sum(r.quantity or 0 for r in today_receipt_rows)

    materials = db.query(Material).filter(Material.user_id == _uid, Material.status == "active").all()
    danger_mats = [m for m in materials if m.safety_stock and m.current_stock < m.safety_stock]
    expiry_warn = db.query(Receipt).filter(
        Receipt.user_id == _uid,
        Receipt.expiry_date != None,
        Receipt.expiry_date <= datetime.now() + timedelta(days=7),
        Receipt.status == "confirmed",
    ).count()

    # 오늘 생산계획 총량
    today_plan_qty = db.query(func.sum(ProductionPlan.planned_quantity)).filter(
        ProductionPlan.user_id == _uid,
        func.date(ProductionPlan.planned_date) == today,
        ProductionPlan.status == "planned",
    ).scalar() or 0

    # 장비 현황
    devices = db.query(Device).filter(Device.user_id == _uid).all()
    device_summary = {
        "total": len(devices),
        "running": sum(1 for d in devices if d.status == "running"),
        "idle": sum(1 for d in devices if d.status == "idle"),
        "error": sum(1 for d in devices if d.status == "error"),
    }

    # 최근 생산실적 4건
    recent_prods_rows = db.query(Production).filter(
        Production.user_id == _uid
    ).order_by(desc(Production.created_at)).limit(4).all()
    recent_productions = []
    for p in recent_prods_rows:
        name = ""
        if p.finished_product and p.finished_product.name:
            name = p.finished_product.name
        elif p.recipe and p.recipe.product_name:
            name = p.recipe.product_name
        recent_productions.append({
            "lot_number": p.lot_number,
            "product_name": name,
            "produced_quantity": p.produced_quantity or 0,
            "good_quantity": p.good_quantity or 0,
            "status": p.status,
            "start_time": p.start_time.strftime("%m/%d %H:%M") if p.start_time else "",
        })

    week_data = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        day_prods = db.query(Production).filter(
            Production.user_id == _uid,
            func.date(Production.start_time) == d,
        ).all()
        good = sum(p.good_quantity or 0 for p in day_prods)
        defect = sum(p.defect_quantity or 0 for p in day_prods)
        week_data.append({
            "label": ["월", "화", "수", "목", "금", "토", "일"][d.weekday()],
            "good": int(good), "defect": int(defect),
        })

    recent_receipts = []
    for r in db.query(Receipt).filter(Receipt.user_id == _uid).order_by(desc(Receipt.created_at)).limit(4).all():
        recent_receipts.append({
            "material_name": r.material.name if r.material else "",
            "supplier_name": r.supplier.name if r.supplier else "",
            "lot_number": r.lot_number,
            "quantity": r.quantity,
            "unit": r.material.unit if r.material else "",
            "created_at": r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
        })

    def _stock_status(m):
        if m.safety_stock and m.current_stock < m.safety_stock:
            return "danger"
        if m.safety_stock and m.current_stock < m.safety_stock * 1.5:
            return "warn"
        return "normal"

    return {
        "today_quantity": int(today_qty),
        "today_good_rate": today_rate,
        "today_receipts": today_receipts,
        "today_receipt_qty": round(today_receipt_qty, 1),
        "yesterday_quantity": int(yesterday_qty),
        "yesterday_good_rate": yesterday_rate,
        "alert_count": len(danger_mats) + expiry_warn,
        "safety_stock_alerts": len(danger_mats),
        "expiry_alerts": expiry_warn,
        "today_plan_qty": int(today_plan_qty),
        "device_summary": device_summary,
        "week_chart": week_data,
        "recent_receipts": recent_receipts,
        "recent_productions": recent_productions,
        "stock_summary": sorted(
            [
                {
                    "name": m.name, "unit": m.unit,
                    "current_stock": m.current_stock or 0,
                    "safety_stock": m.safety_stock or 0,
                    "status": _stock_status(m),
                }
                for m in materials
            ],
            key=lambda x: {"danger": 0, "warn": 1, "normal": 2}[x["status"]]
        ),
    }


# ─────────────────────────────────────────
# Device (장비) — 등록 · 관리 · 통계
# ─────────────────────────────────────────
class DeviceIn(BaseModel):
    name: Optional[str] = None
    device_code: Optional[str] = None
    process_id: Optional[int] = None
    status: Optional[str] = None
    photo_data: Optional[str] = None
    maintenance_notes: Optional[str] = None
    collect_production: Optional[bool] = None


def _device_period_range(period: str):
    today = datetime.now().date()
    now = datetime.now()
    if period == "week":
        week_start = today - timedelta(days=today.weekday())
        return datetime.combine(week_start, datetime.min.time()), now
    elif period == "month":
        return datetime(today.year, today.month, 1), now
    else:  # day
        return datetime.combine(today, datetime.min.time()), now


@app.get("/api/devices")
def list_devices(period: str = "day", db: Session = Depends(get_db)):
    items = db.query(Device).filter(Device.user_id == uid()).order_by(Device.created_at).all()
    start_dt, end_dt = _device_period_range(period)
    today = datetime.now().date()
    result = []
    for d in items:
        prods = db.query(Production).filter(
            Production.user_id == uid(),
            Production.device_id == d.id,
            Production.start_time >= start_dt,
            Production.start_time <= end_dt,
        ).all()
        period_qty = sum(p.produced_quantity or 0 for p in prods)
        period_hours = sum(
            max(0, (p.end_time - p.start_time).total_seconds() / 3600)
            for p in prods if p.start_time and p.end_time
        )
        today_count = db.query(Production).filter(
            Production.device_id == d.id,
            func.date(Production.created_at) == today,
        ).count()
        result.append({
            "id": d.id, "device_code": d.device_code, "name": d.name,
            "process_id": d.process_id,
            "process_name": d.process.name if d.process else "",
            "status": d.status,
            "collect_production": bool(d.collect_production),
            "period_quantity": int(period_qty),
            "period_hours": round(period_hours, 1),
            "today_count": today_count,
            "last_received_at": d.last_received_at.strftime("%Y-%m-%d %H:%M") if d.last_received_at else None,
        })
    return result


@app.get("/api/devices/{device_id}")
def get_device(device_id: int, db: Session = Depends(get_db)):
    d = db.query(Device).filter(Device.id == device_id, Device.user_id == uid()).first()
    if not d:
        raise HTTPException(404, "Not found")
    return {
        "id": d.id, "device_code": d.device_code, "name": d.name,
        "process_id": d.process_id,
        "process_name": d.process.name if d.process else "",
        "status": d.status,
        "photo_data": d.photo_data,
        "maintenance_notes": d.maintenance_notes or "",
        "last_received_at": d.last_received_at.strftime("%Y-%m-%d %H:%M") if d.last_received_at else None,
    }


@app.post("/api/devices", status_code=201)
def create_device(data: DeviceIn, db: Session = Depends(get_db)):
    if not data.name:
        raise HTTPException(422, "name is required")
    now = datetime.now()
    code = data.device_code or f"DEV-{now.strftime('%Y%m%d')}-{secrets.token_hex(2).upper()}"
    d = Device(
        device_code=code,
        name=data.name,
        process_id=data.process_id,
        status=data.status or "idle",
        photo_data=data.photo_data,
        maintenance_notes=data.maintenance_notes,
        api_key=f"fsk_{secrets.token_hex(16)}",
        user_id=uid(),
    )
    db.add(d); db.commit(); db.refresh(d)
    return {"id": d.id, "device_code": d.device_code}


@app.put("/api/devices/{device_id}")
def update_device(device_id: int, data: DeviceIn, db: Session = Depends(get_db)):
    d = db.query(Device).filter(Device.id == device_id, Device.user_id == uid()).first()
    if not d:
        raise HTTPException(404, "Not found")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(d, k, v)
    db.commit()
    return {"ok": True}


@app.delete("/api/devices/{device_id}")
def delete_device(device_id: int, db: Session = Depends(get_db)):
    d = db.query(Device).filter(Device.id == device_id, Device.user_id == uid()).first()
    if not d:
        raise HTTPException(404, "Not found")
    db.delete(d); db.commit()
    return {"ok": True}


@app.get("/api/devices/{device_id}/stats")
def device_stats(device_id: int, period: str = "day", db: Session = Depends(get_db)):
    d = db.query(Device).filter(Device.id == device_id, Device.user_id == uid()).first()
    if not d:
        raise HTTPException(404, "Not found")
    today = datetime.now().date()
    entries = []
    if period == "day":
        for i in range(6, -1, -1):
            dt = today - timedelta(days=i)
            prods = db.query(Production).filter(
                Production.device_id == device_id,
                func.date(Production.start_time) == dt,
            ).all()
            qty = int(sum(p.produced_quantity or 0 for p in prods))
            hrs = round(sum(
                max(0, (p.end_time - p.start_time).total_seconds() / 3600)
                for p in prods if p.start_time and p.end_time
            ), 1)
            entries.append({"label": dt.strftime("%m/%d"), "quantity": qty, "hours": hrs})
    elif period == "week":
        week_start = today - timedelta(days=today.weekday())
        for i in range(7, -1, -1):
            ws = week_start - timedelta(weeks=i)
            we = ws + timedelta(days=6)
            prods = db.query(Production).filter(
                Production.device_id == device_id,
                func.date(Production.start_time) >= ws,
                func.date(Production.start_time) <= we,
            ).all()
            qty = int(sum(p.produced_quantity or 0 for p in prods))
            hrs = round(sum(
                max(0, (p.end_time - p.start_time).total_seconds() / 3600)
                for p in prods if p.start_time and p.end_time
            ), 1)
            entries.append({"label": ws.strftime("%m/%d") + "주", "quantity": qty, "hours": hrs})
    else:  # month
        now = datetime.now()
        for i in range(5, -1, -1):
            m = now.month - i
            y = now.year
            while m <= 0:
                m += 12; y -= 1
            prods = db.query(Production).filter(
                Production.device_id == device_id,
                func.strftime('%Y', Production.start_time) == str(y),
                func.strftime('%m', Production.start_time) == f"{m:02d}",
            ).all()
            qty = int(sum(p.produced_quantity or 0 for p in prods))
            hrs = round(sum(
                max(0, (p.end_time - p.start_time).total_seconds() / 3600)
                for p in prods if p.start_time and p.end_time
            ), 1)
            entries.append({"label": f"{m}월", "quantity": qty, "hours": hrs})
    return {"device_id": device_id, "period": period, "entries": entries}


# ─────────────────────────────────────────
# 장비 생산량 수집 & 제품 배분
# ─────────────────────────────────────────
@app.get("/api/device-productions/today")
def device_productions_today(db: Session = Depends(get_db)):
    today = datetime.now().date()
    devices = db.query(Device).filter(Device.user_id == uid(), Device.collect_production == True).order_by(Device.id).all()
    result = []
    for d in devices:
        # 오늘 장비 API 수신 수량 (device / device_raw)
        device_prods = db.query(Production).filter(
            Production.user_id == uid(),
            Production.device_id == d.id,
            Production.input_method == "device",
            func.date(Production.start_time) == today,
        ).all()
        collected_qty = int(sum(p.produced_quantity or 0 for p in device_prods))

        # 오늘 배분 완료된 records
        alloc_prods = db.query(Production).filter(
            Production.user_id == uid(),
            Production.device_id == d.id,
            Production.input_method == "device_allocated",
            func.date(Production.start_time) == today,
        ).all()

        allocations = []
        for ap in alloc_prods:
            allocations.append({
                "id": ap.id,
                "product_id": ap.finished_product_id,
                "product_name": ap.finished_product.name if ap.finished_product else "기타",
                "product_unit": ap.finished_product.unit if ap.finished_product else "개",
                "quantity": ap.produced_quantity or 0,
                "note": ap.note or "",
            })

        result.append({
            "device_id": d.id,
            "device_name": d.name,
            "device_code": d.device_code,
            "collected_qty": collected_qty,
            "is_allocated": len(alloc_prods) > 0,
            "allocations": allocations,
        })
    return result


class DeviceProductionSave(BaseModel):
    device_id: int
    total_quantity: float
    production_date: Optional[str] = None
    allocations: List[dict]  # [{product_id, quantity, note}]


@app.post("/api/device-productions/save", status_code=200)
def save_device_production(data: DeviceProductionSave, db: Session = Depends(get_db)):
    if data.production_date:
        prod_dt = datetime.fromisoformat(data.production_date)
    else:
        prod_dt = datetime.now()
    prod_date = prod_dt.date()

    _uid = uid()
    # 기존 device_allocated 레코드 삭제 + 재고 복원
    existing = db.query(Production).filter(
        Production.user_id == _uid,
        Production.device_id == data.device_id,
        Production.input_method == "device_allocated",
        func.date(Production.start_time) == prod_date,
    ).all()
    for ep in existing:
        if ep.finished_product_id and ep.produced_quantity:
            for mat_id, amt in explode_bom(ep.finished_product_id, ep.produced_quantity, db).items():
                mat = db.query(Material).filter(Material.id == mat_id, Material.user_id == _uid).first()
                if mat:
                    mat.current_stock = mat.current_stock + round(amt, 3)
        db.delete(ep)
    db.flush()

    created = 0
    for alloc in data.allocations:
        qty = float(alloc.get("quantity", 0) or 0)
        if qty <= 0:
            continue
        pid = alloc.get("product_id") or None
        note = (alloc.get("note") or "").strip() or ("기타" if not pid else "")

        lot = f"DA-{prod_dt.strftime('%Y%m%d')}-{secrets.token_hex(2).upper()}"
        prod = Production(
            lot_number=lot,
            finished_product_id=pid,
            device_id=data.device_id,
            produced_quantity=qty,
            good_quantity=qty,
            defect_quantity=0,
            start_time=prod_dt,
            end_time=prod_dt,
            input_method="device_allocated",
            status="completed",
            note=note,
            user_id=_uid,
        )
        db.add(prod)

        if pid and qty:
            for mat_id, amt in explode_bom(pid, qty, db).items():
                mat = db.query(Material).filter(Material.id == mat_id, Material.user_id == _uid).first()
                if mat:
                    mat.current_stock = max(0, mat.current_stock - round(amt, 3))
        created += 1

    db.commit()
    return {"ok": True, "created": created}


@app.post("/api/equipment/production")
def device_production(
    data: DeviceProductionIn,
    db: Session = Depends(get_db)
):
    device = db.query(Device).filter(Device.device_code == data.device_id).first()
    if not device:
        raise HTTPException(404, "Device not found")
    recipe = db.query(Recipe).filter(Recipe.product_code == data.product_code).first()
    if not recipe:
        raise HTTPException(404, "Recipe not found")

    total_defects = sum(d.get("quantity", 0) for d in (data.defects or []))
    now = datetime.now()
    lot = f"PROD-{now.strftime('%Y%m%d%H%M%S')}-{device.device_code}"

    prod = Production(
        lot_number=lot, recipe_id=recipe.id, device_id=device.id,
        produced_quantity=data.produced_quantity,
        good_quantity=data.produced_quantity - total_defects,
        defect_quantity=total_defects,
        start_time=now, end_time=now,
        input_method="device", status="completed",
    )
    db.add(prod)
    device.status = "running"
    device.last_received_at = now

    if recipe.base_quantity:
        ratio = data.produced_quantity / recipe.base_quantity
        for ing in recipe.ingredients:
            mat = db.query(Material).get(ing.material_id)
            if mat:
                mat.current_stock = max(0, mat.current_stock - ing.quantity * ratio)

    db.commit()
    return {"ok": True, "lot_number": lot}


# ═══════════════════════════════════════════════════════
# 기본정보 마스터 데이터 CRUD
# ═══════════════════════════════════════════════════════

# ── Pydantic schemas ──────────────────────────────────
class MaterialMasterIn(BaseModel):
    material_code: Optional[str] = None
    name: str
    category: Optional[str] = None
    unit: str = "kg"
    safety_stock: float = 0
    unit_price: Optional[float] = None
    description: Optional[str] = None
    status: str = "active"

class SemiProductIn(BaseModel):
    code: Optional[str] = None
    name: str
    category: Optional[str] = None
    unit: str = "kg"
    standard_qty: Optional[float] = None
    unit_price: Optional[float] = None
    description: Optional[str] = None
    status: str = "active"
    unit_conv_qty: Optional[float] = None
    unit_conv_unit: Optional[str] = None
    unit_conv2_qty: Optional[float] = None
    unit_conv2_unit: Optional[str] = None

class FinishedProductIn(BaseModel):
    code: Optional[str] = None
    name: str
    category: Optional[str] = None
    unit: str = "ea"
    unit_price: Optional[float] = None
    description: Optional[str] = None
    status: str = "active"
    unit_conv_qty: Optional[float] = None
    unit_conv_unit: Optional[str] = None
    unit_conv2_qty: Optional[float] = None
    unit_conv2_unit: Optional[str] = None

class PartnerIn(BaseModel):
    name: str
    business_number: Optional[str] = None
    partner_type: str = "supplier"  # supplier / customer / other
    contact_person: Optional[str] = None
    contact: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    main_products: Optional[str] = None
    status: str = "active"

class SalesLogIn(BaseModel):
    log_date: str
    content: str
    author: Optional[str] = None

class SemiProductBOMIn(BaseModel):
    material_id: Optional[int] = None
    sub_semi_id: Optional[int] = None
    quantity: float
    unit: str = "kg"
    note: Optional[str] = None

class SemiProductProcessIn(BaseModel):
    process_id: int
    step_order: int = 1
    note: Optional[str] = None


# ── 원재료 마스터 ─────────────────────────────────────
@app.get("/api/master/materials")
def master_list_materials(
    search: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    q = db.query(Material).filter(Material.user_id == uid())
    if search:
        q = q.filter(Material.name.contains(search))
    if category:
        q = q.filter(Material.category == category)
    if status:
        q = q.filter(Material.status == status)
    items = q.order_by(Material.id).all()
    result = []
    for m in items:
        s = "normal"
        if m.current_stock <= 0:
            s = "danger"
        elif m.current_stock < m.safety_stock:
            s = "danger"
        elif m.current_stock < m.safety_stock * 1.5:
            s = "warn"
        result.append({
            "id": m.id,
            "material_code": m.material_code or f"MAT-{m.id:03d}",
            "name": m.name, "category": m.category, "unit": m.unit,
            "safety_stock": m.safety_stock, "current_stock": m.current_stock,
            "unit_price": m.unit_price, "description": m.description,
            "status": m.status or "active", "stock_status": s,
        })
    return result

@app.post("/api/master/materials", status_code=201)
def master_create_material(data: MaterialMasterIn, db: Session = Depends(get_db)):
    m = Material(**data.model_dump(), user_id=uid())
    db.add(m); db.commit(); db.refresh(m)
    return {"id": m.id}

@app.put("/api/master/materials/{mid}")
def master_update_material(mid: int, data: MaterialMasterIn, db: Session = Depends(get_db)):
    m = db.query(Material).filter(Material.id == mid, Material.user_id == uid()).first()
    if not m:
        raise HTTPException(404, "Not found")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(m, k, v)
    db.commit()
    return {"ok": True}

@app.delete("/api/master/materials/{mid}")
def master_delete_material(mid: int, db: Session = Depends(get_db)):
    m = db.query(Material).filter(Material.id == mid, Material.user_id == uid()).first()
    if not m:
        raise HTTPException(404, "Not found")
    m.status = "inactive"
    db.commit()
    return {"ok": True}


# ── 반제품 마스터 ─────────────────────────────────────
@app.get("/api/master/semi-products")
def master_list_semi(
    search: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    q = db.query(SemiProduct).filter(SemiProduct.user_id == uid())
    if search:
        q = q.filter(SemiProduct.name.contains(search))
    if status:
        q = q.filter(SemiProduct.status == status)
    return [
        {"id": s.id, "code": s.code or f"SEM-{s.id:03d}", "name": s.name,
         "category": s.category, "unit": s.unit, "standard_qty": s.standard_qty,
         "unit_price": s.unit_price, "description": s.description, "status": s.status,
         "unit_conv_qty": s.unit_conv_qty, "unit_conv_unit": s.unit_conv_unit,
         "unit_conv2_qty": s.unit_conv2_qty, "unit_conv2_unit": s.unit_conv2_unit,
         "created_at": s.created_at.strftime("%Y-%m-%d") if s.created_at else ""}
        for s in q.order_by(SemiProduct.id).all()
    ]

@app.post("/api/master/semi-products", status_code=201)
def master_create_semi(data: SemiProductIn, db: Session = Depends(get_db)):
    s = SemiProduct(**data.model_dump(), user_id=uid())
    db.add(s); db.commit(); db.refresh(s)
    return {"id": s.id}

@app.put("/api/master/semi-products/{sid}")
def master_update_semi(sid: int, data: SemiProductIn, db: Session = Depends(get_db)):
    s = db.query(SemiProduct).filter(SemiProduct.id == sid, SemiProduct.user_id == uid()).first()
    if not s:
        raise HTTPException(404, "Not found")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(s, k, v)
    db.commit()
    return {"ok": True}

@app.delete("/api/master/semi-products/{sid}")
def master_delete_semi(sid: int, db: Session = Depends(get_db)):
    s = db.query(SemiProduct).filter(SemiProduct.id == sid, SemiProduct.user_id == uid()).first()
    if not s:
        raise HTTPException(404, "Not found")
    s.status = "inactive"
    db.commit()
    return {"ok": True}


# ── 제품 마스터 ───────────────────────────────────────
@app.get("/api/master/products")
def master_list_products(
    search: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    q = db.query(FinishedProduct).filter(FinishedProduct.user_id == uid())
    if search:
        q = q.filter(FinishedProduct.name.contains(search))
    if status:
        q = q.filter(FinishedProduct.status == status)
    return [
        {"id": p.id, "code": p.code or f"PRD-{p.id:03d}", "name": p.name,
         "category": p.category, "unit": p.unit,
         "unit_price": p.unit_price, "current_stock": p.current_stock or 0,
         "description": p.description, "status": p.status,
         "unit_conv_qty": p.unit_conv_qty, "unit_conv_unit": p.unit_conv_unit,
         "unit_conv2_qty": p.unit_conv2_qty, "unit_conv2_unit": p.unit_conv2_unit,
         "created_at": p.created_at.strftime("%Y-%m-%d") if p.created_at else ""}
        for p in q.order_by(FinishedProduct.id).all()
    ]

@app.post("/api/master/products", status_code=201)
def master_create_product(data: FinishedProductIn, db: Session = Depends(get_db)):
    p = FinishedProduct(**data.model_dump(), user_id=uid())
    db.add(p); db.commit(); db.refresh(p)
    return {"id": p.id}

@app.put("/api/master/products/{pid}")
def master_update_product(pid: int, data: FinishedProductIn, db: Session = Depends(get_db)):
    p = db.query(FinishedProduct).filter(FinishedProduct.id == pid, FinishedProduct.user_id == uid()).first()
    if not p:
        raise HTTPException(404, "Not found")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(p, k, v)
    db.commit()
    return {"ok": True}

@app.delete("/api/master/products/{pid}")
def master_delete_product(pid: int, db: Session = Depends(get_db)):
    p = db.query(FinishedProduct).filter(FinishedProduct.id == pid, FinishedProduct.user_id == uid()).first()
    if not p:
        raise HTTPException(404, "Not found")
    p.status = "inactive"
    db.commit()
    return {"ok": True}


# ── 거래처 마스터 ─────────────────────────────────────
@app.get("/api/master/partners")
def master_list_partners(
    search: Optional[str] = None,
    partner_type: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    q = db.query(Supplier).filter(Supplier.user_id == uid())
    if search:
        q = q.filter(Supplier.name.contains(search) | Supplier.business_number.contains(search))
    if partner_type:
        q = q.filter(Supplier.partner_type == partner_type)
    if status:
        q = q.filter(Supplier.status == status)
    return [
        {"id": s.id, "name": s.name, "business_number": s.business_number,
         "partner_type": s.partner_type or "supplier",
         "contact_person": s.contact_person, "contact": s.contact,
         "email": s.email, "address": s.address,
         "main_products": s.main_products,
         "status": s.status, "ocr_mapped": s.ocr_mapped,
         "log_count": len(s.sales_logs),
         "created_at": s.created_at.strftime("%Y-%m-%d") if s.created_at else ""}
        for s in q.order_by(Supplier.id).all()
    ]

@app.post("/api/master/partners", status_code=201)
def master_create_partner(data: PartnerIn, db: Session = Depends(get_db)):
    s = Supplier(**data.model_dump(), user_id=uid())
    db.add(s); db.commit(); db.refresh(s)
    return {"id": s.id}

@app.put("/api/master/partners/{pid}")
def master_update_partner(pid: int, data: PartnerIn, db: Session = Depends(get_db)):
    s = db.query(Supplier).filter(Supplier.id == pid, Supplier.user_id == uid()).first()
    if not s:
        raise HTTPException(404, "Not found")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(s, k, v)
    db.commit()
    return {"ok": True}

@app.delete("/api/master/partners/{pid}")
def master_delete_partner(pid: int, db: Session = Depends(get_db)):
    s = db.query(Supplier).filter(Supplier.id == pid, Supplier.user_id == uid()).first()
    if not s:
        raise HTTPException(404, "Not found")
    s.status = "inactive"
    db.commit()
    return {"ok": True}


# ── 반제품 BOM ─────────────────────────────────────────
@app.get("/api/master/semi-products/{sid}/bom")
def get_semi_bom(sid: int, db: Session = Depends(get_db)):
    items = db.query(SemiProductBOM).filter(SemiProductBOM.semi_product_id == sid).all()
    return [
        {"id": b.id,
         "material_id": b.material_id,
         "material_name": b.material.name if b.material else None,
         "material_unit": b.material.unit if b.material else None,
         "sub_semi_id": b.sub_semi_id,
         "sub_semi_name": b.sub_semi.name if b.sub_semi else None,
         "quantity": b.quantity, "unit": b.unit, "note": b.note,
         "item_type": "material" if b.material_id else "semi"}
        for b in items
    ]

@app.post("/api/master/semi-products/{sid}/bom", status_code=201)
def add_semi_bom(sid: int, data: SemiProductBOMIn, db: Session = Depends(get_db)):
    if not db.query(SemiProduct).get(sid):
        raise HTTPException(404, "SemiProduct not found")
    b = SemiProductBOM(semi_product_id=sid, **data.model_dump())
    db.add(b); db.commit(); db.refresh(b)
    return {"id": b.id}

@app.delete("/api/master/semi-products/{sid}/bom/{bid}")
def delete_semi_bom(sid: int, bid: int, db: Session = Depends(get_db)):
    b = db.query(SemiProductBOM).filter(SemiProductBOM.id == bid, SemiProductBOM.semi_product_id == sid).first()
    if not b:
        raise HTTPException(404, "Not found")
    db.delete(b); db.commit()
    return {"ok": True}


# ── 반제품 공정 ────────────────────────────────────────
@app.get("/api/master/semi-products/{sid}/processes")
def get_semi_processes(sid: int, db: Session = Depends(get_db)):
    items = db.query(SemiProductProcess).filter(
        SemiProductProcess.semi_product_id == sid
    ).order_by(SemiProductProcess.step_order).all()
    return [
        {"id": pp.id, "process_id": pp.process_id,
         "process_name": pp.process.name if pp.process else "",
         "process_code": pp.process.code if pp.process else "",
         "step_order": pp.step_order, "note": pp.note}
        for pp in items
    ]

@app.post("/api/master/semi-products/{sid}/processes", status_code=201)
def add_semi_process(sid: int, data: SemiProductProcessIn, db: Session = Depends(get_db)):
    if not db.query(SemiProduct).get(sid):
        raise HTTPException(404, "SemiProduct not found")
    pp = SemiProductProcess(semi_product_id=sid, **data.model_dump())
    db.add(pp); db.commit(); db.refresh(pp)
    return {"id": pp.id}

@app.delete("/api/master/semi-products/{sid}/processes/{ppid}")
def delete_semi_process(sid: int, ppid: int, db: Session = Depends(get_db)):
    pp = db.query(SemiProductProcess).filter(
        SemiProductProcess.id == ppid, SemiProductProcess.semi_product_id == sid
    ).first()
    if not pp:
        raise HTTPException(404, "Not found")
    db.delete(pp); db.commit()
    return {"ok": True}


# ── 영업일지 ──────────────────────────────────────────
@app.get("/api/master/partners/{pid}/sales-logs")
def get_sales_logs(pid: int, db: Session = Depends(get_db)):
    logs = db.query(SalesLog).filter(SalesLog.supplier_id == pid).order_by(desc(SalesLog.log_date)).all()
    return [
        {"id": lg.id, "log_date": lg.log_date.strftime("%Y-%m-%d"),
         "content": lg.content, "author": lg.author,
         "created_at": lg.created_at.strftime("%Y-%m-%d %H:%M") if lg.created_at else ""}
        for lg in logs
    ]

@app.post("/api/master/partners/{pid}/sales-logs", status_code=201)
def create_sales_log(pid: int, data: SalesLogIn, db: Session = Depends(get_db)):
    if not db.query(Supplier).get(pid):
        raise HTTPException(404, "Partner not found")
    lg = SalesLog(
        supplier_id=pid,
        log_date=datetime.fromisoformat(data.log_date),
        content=data.content,
        author=data.author,
    )
    db.add(lg); db.commit(); db.refresh(lg)
    return {"id": lg.id}

@app.delete("/api/master/partners/{pid}/sales-logs/{lid}")
def delete_sales_log(pid: int, lid: int, db: Session = Depends(get_db)):
    lg = db.query(SalesLog).filter(SalesLog.id == lid, SalesLog.supplier_id == pid).first()
    if not lg:
        raise HTTPException(404, "Not found")
    db.delete(lg); db.commit()
    return {"ok": True}


# ── 원재료-거래처 연결 ────────────────────────────────
@app.get("/api/master/materials/{mid}/suppliers")
def get_material_suppliers(mid: int, db: Session = Depends(get_db)):
    links = db.query(MaterialSupplier).filter(MaterialSupplier.material_id == mid).all()
    return [{"id": lk.id, "supplier_id": lk.supplier_id,
             "supplier_name": lk.supplier.name if lk.supplier else "",
             "is_primary": lk.is_primary} for lk in links]

class MaterialSupplierIn(BaseModel):
    supplier_ids: List[int]
    primary_id: Optional[int] = None

@app.put("/api/master/materials/{mid}/suppliers")
def set_material_suppliers(mid: int, data: MaterialSupplierIn, db: Session = Depends(get_db)):
    m = db.query(Material).get(mid)
    if not m:
        raise HTTPException(404, "Not found")
    db.query(MaterialSupplier).filter(MaterialSupplier.material_id == mid).delete()
    for sid in data.supplier_ids:
        lk = MaterialSupplier(
            material_id=mid,
            supplier_id=sid,
            is_primary=(sid == data.primary_id),
        )
        db.add(lk)
    db.commit()
    return {"ok": True}


# ═══════════════════════════════════════════════════════
# 공정 마스터 CRUD
# ═══════════════════════════════════════════════════════
class ProcessIn(BaseModel):
    code: Optional[str] = None
    name: str
    description: Optional[str] = None
    status: str = "active"

@app.get("/api/master/processes")
def master_list_processes(
    search: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    q = db.query(Process).filter(Process.user_id == uid())
    if search:
        q = q.filter(Process.name.contains(search))
    if status:
        q = q.filter(Process.status == status)
    return [
        {"id": p.id, "code": p.code or f"PRC-{p.id:03d}", "name": p.name,
         "description": p.description, "status": p.status,
         "created_at": p.created_at.strftime("%Y-%m-%d") if p.created_at else ""}
        for p in q.order_by(Process.id).all()
    ]

@app.post("/api/master/processes", status_code=201)
def master_create_process(data: ProcessIn, db: Session = Depends(get_db)):
    p = Process(**data.model_dump(), user_id=uid())
    db.add(p); db.commit(); db.refresh(p)
    return {"id": p.id}

@app.put("/api/master/processes/{pid}")
def master_update_process(pid: int, data: ProcessIn, db: Session = Depends(get_db)):
    p = db.query(Process).filter(Process.id == pid, Process.user_id == uid()).first()
    if not p:
        raise HTTPException(404, "Not found")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(p, k, v)
    db.commit()
    return {"ok": True}

@app.delete("/api/master/processes/{pid}")
def master_delete_process(pid: int, db: Session = Depends(get_db)):
    p = db.query(Process).filter(Process.id == pid, Process.user_id == uid()).first()
    if not p:
        raise HTTPException(404, "Not found")
    p.status = "inactive"
    db.commit()
    return {"ok": True}


# ═══════════════════════════════════════════════════════
# 엑셀 템플릿 다운로드 / 일괄 업로드
# ═══════════════════════════════════════════════════════

_EXCEL_HEADERS = {
    "material": ["품목명*", "코드", "분류", "단위*", "안전재고", "단가(원)", "설명"],
    "semi":     ["품목명*", "코드", "분류", "단위*", "표준생산량", "단가(원)", "설명"],
    "product":  ["제품명*", "코드", "분류", "단위*", "단가(원)", "설명"],
    "partner":  ["업체명*", "사업자번호", "구분(supplier/customer/other)", "담당자", "연락처", "이메일", "주소", "주요제품"],
    "process":  ["공정명*", "코드", "설명"],
}

_EXCEL_SAMPLES = {
    "material": [["밀가루", "MAT-001", "곡물류", "kg", 50, 1200, "강력분"]],
    "semi":     [["반죽A", "SEM-001", "", "kg", 100, 800, ""]],
    "product":  [["식빵 500g", "PRD-001", "", "ea", 3500, ""]],
    "partner":  [["(주)베스트푸드", "123-45-67890", "supplier", "김담당", "010-1234-5678", "", "", "밀가루, 설탕"]],
    "process":  [["반죽", "PRC-001", "재료 혼합 및 반죽"]],
}

_EXCEL_TITLES = {
    "material": "원재료", "semi": "반제품", "product": "제품",
    "partner": "거래처", "process": "공정",
}

def _make_excel_template(dtype: str) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _EXCEL_TITLES.get(dtype, dtype)

    header_fill = PatternFill("solid", fgColor="1B7A5E")
    req_fill    = PatternFill("solid", fgColor="FF6B6B")
    sample_fill = PatternFill("solid", fgColor="F0FFF8")
    hdr_font    = Font(bold=True, color="FFFFFF", size=11)
    border_side = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    headers = _EXCEL_HEADERS[dtype]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = req_fill if h.endswith("*") else header_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(len(h) * 2.2, 14)

    for row_data in _EXCEL_SAMPLES[dtype]:
        ws.append(row_data)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = sample_fill
            cell.border = thin_border

    # 안내 행
    ws.append([])
    note_row = ws.max_row + 1
    ws.cell(row=note_row, column=1, value="※ * 표시 항목은 필수입력입니다. 2행은 예시이므로 삭제 후 입력하세요.")
    ws.cell(row=note_row, column=1).font = Font(color="888888", italic=True, size=9)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@app.get("/api/master/excel-template/{dtype}")
def download_excel_template(dtype: str):
    if dtype not in _EXCEL_HEADERS:
        raise HTTPException(400, "지원하지 않는 유형입니다")
    buf = _make_excel_template(dtype)
    filename = f"{_EXCEL_TITLES[dtype]}_양식.xlsx"
    encoded = quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )

@app.post("/api/master/excel-import/{dtype}", status_code=200)
async def import_excel(dtype: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    if dtype not in _EXCEL_HEADERS:
        raise HTTPException(400, "지원하지 않는 유형입니다")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(400, "엑셀 파일을 읽을 수 없습니다")

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        if not row or all(v is None for v in row):
            continue
        def v(idx): return str(row[idx]).strip() if row[idx] is not None else ""
        def n(idx):
            try: return float(row[idx]) if row[idx] is not None else None
            except: return None

        try:
            if dtype == "material":
                if not v(0): errors.append(f"{i}행: 품목명 누락"); continue
                db.add(Material(name=v(0), material_code=v(1) or None, category=v(2) or None,
                    unit=v(3) or "kg", safety_stock=n(4) or 0, unit_price=n(5),
                    description=v(6) or None, status="active", user_id=uid()))
            elif dtype == "semi":
                if not v(0): errors.append(f"{i}행: 품목명 누락"); continue
                db.add(SemiProduct(name=v(0), code=v(1) or None, category=v(2) or None,
                    unit=v(3) or "kg", standard_qty=n(4), unit_price=n(5),
                    description=v(6) or None, status="active", user_id=uid()))
            elif dtype == "product":
                if not v(0): errors.append(f"{i}행: 제품명 누락"); continue
                db.add(FinishedProduct(name=v(0), code=v(1) or None, category=v(2) or None,
                    unit=v(3) or "ea", unit_price=n(4),
                    description=v(5) or None, status="active", user_id=uid()))
            elif dtype == "partner":
                if not v(0): errors.append(f"{i}행: 업체명 누락"); continue
                ptype = v(2) if v(2) in ("supplier","customer","other") else "supplier"
                db.add(Supplier(name=v(0), business_number=v(1) or None, partner_type=ptype,
                    contact_person=v(3) or None, contact=v(4) or None, email=v(5) or None,
                    address=v(6) or None, main_products=v(7) or None, status="active", user_id=uid()))
            elif dtype == "process":
                if not v(0): errors.append(f"{i}행: 공정명 누락"); continue
                db.add(Process(name=v(0), code=v(1) or None, description=v(2) or None,
                    status="active", user_id=uid()))
            created += 1
        except Exception as e:
            errors.append(f"{i}행 오류: {str(e)}")

    db.commit()
    return {"created": created, "errors": errors}


# ═══════════════════════════════════════════════════════
# 제품 BOM CRUD
# ═══════════════════════════════════════════════════════
class BOMItemIn(BaseModel):
    material_id: Optional[int] = None
    semi_product_id: Optional[int] = None
    child_product_id: Optional[int] = None
    quantity: float
    unit: str = "kg"
    note: Optional[str] = None

@app.get("/api/master/products/{pid}/bom")
def get_product_bom(pid: int, db: Session = Depends(get_db)):
    items = db.query(ProductBOM).filter(ProductBOM.product_id == pid).all()
    result = []
    for b in items:
        if b.material_id:
            item_type = "material"
        elif b.semi_product_id:
            item_type = "semi"
        else:
            item_type = "product"
        result.append({
            "id": b.id,
            "material_id": b.material_id,
            "material_name": b.material.name if b.material else None,
            "material_unit": b.material.unit if b.material else None,
            "semi_product_id": b.semi_product_id,
            "semi_product_name": b.semi_product.name if b.semi_product else None,
            "child_product_id": b.child_product_id,
            "child_product_name": b.child_product.name if b.child_product else None,
            "child_product_unit": b.child_product.unit if b.child_product else None,
            "quantity": b.quantity,
            "unit": b.unit,
            "note": b.note,
            "item_type": item_type,
        })
    return result

@app.post("/api/master/products/{pid}/bom", status_code=201)
def add_bom_item(pid: int, data: BOMItemIn, db: Session = Depends(get_db)):
    if not db.query(FinishedProduct).get(pid):
        raise HTTPException(404, "Product not found")
    b = ProductBOM(product_id=pid, **data.model_dump())
    db.add(b); db.commit(); db.refresh(b)
    return {"id": b.id}

@app.put("/api/master/products/{pid}/bom/{bid}")
def update_bom_item(pid: int, bid: int, data: BOMItemIn, db: Session = Depends(get_db)):
    b = db.query(ProductBOM).filter(ProductBOM.id == bid, ProductBOM.product_id == pid).first()
    if not b:
        raise HTTPException(404, "Not found")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(b, k, v)
    db.commit()
    return {"ok": True}

@app.delete("/api/master/products/{pid}/bom/{bid}")
def delete_bom_item(pid: int, bid: int, db: Session = Depends(get_db)):
    b = db.query(ProductBOM).filter(ProductBOM.id == bid, ProductBOM.product_id == pid).first()
    if not b:
        raise HTTPException(404, "Not found")
    db.delete(b); db.commit()
    return {"ok": True}


# ═══════════════════════════════════════════════════════
# 제품 공정 CRUD
# ═══════════════════════════════════════════════════════
class ProductProcessIn(BaseModel):
    process_id: int
    step_order: int = 1
    note: Optional[str] = None

@app.get("/api/master/products/{pid}/processes")
def get_product_processes(pid: int, db: Session = Depends(get_db)):
    items = db.query(ProductProcess).filter(
        ProductProcess.product_id == pid
    ).order_by(ProductProcess.step_order).all()
    return [
        {"id": pp.id, "process_id": pp.process_id,
         "process_name": pp.process.name if pp.process else "",
         "process_code": pp.process.code if pp.process else "",
         "step_order": pp.step_order, "note": pp.note}
        for pp in items
    ]

@app.post("/api/master/products/{pid}/processes", status_code=201)
def add_product_process(pid: int, data: ProductProcessIn, db: Session = Depends(get_db)):
    if not db.query(FinishedProduct).get(pid):
        raise HTTPException(404, "Product not found")
    pp = ProductProcess(product_id=pid, **data.model_dump())
    db.add(pp); db.commit(); db.refresh(pp)
    return {"id": pp.id}

@app.delete("/api/master/products/{pid}/processes/{ppid}")
def delete_product_process(pid: int, ppid: int, db: Session = Depends(get_db)):
    pp = db.query(ProductProcess).filter(
        ProductProcess.id == ppid, ProductProcess.product_id == pid
    ).first()
    if not pp:
        raise HTTPException(404, "Not found")
    db.delete(pp); db.commit()
    return {"ok": True}

@app.put("/api/master/products/{pid}/processes/reorder")
def reorder_product_processes(pid: int, order: List[int], db: Session = Depends(get_db)):
    """order: list of ProductProcess IDs in desired order"""
    for idx, ppid in enumerate(order, start=1):
        pp = db.query(ProductProcess).filter(
            ProductProcess.id == ppid, ProductProcess.product_id == pid
        ).first()
        if pp:
            pp.step_order = idx
    db.commit()
    return {"ok": True}


# ═══════════════════════════════════════════════════════
# 생산계획 CRUD
# ═══════════════════════════════════════════════════════

@app.get("/api/production-plans/bom-preview")
def plan_bom_preview(product_id: int, quantity: float, db: Session = Depends(get_db)):
    """생산계획 입력 시 BOM 기반 원재료 소요량 + 현재 재고 비교"""
    product = db.query(FinishedProduct).filter(
        FinishedProduct.id == product_id,
        FinishedProduct.user_id == uid()
    ).first()
    if not product:
        raise HTTPException(404, "Product not found")

    mat_reqs = explode_bom(product_id, quantity, db)  # {material_id: required_qty}

    result = []
    for mat_id, req_qty in mat_reqs.items():
        mat = db.query(Material).filter(
            Material.id == mat_id,
            Material.user_id == uid()
        ).first()
        if not mat:
            continue
        stock = mat.current_stock or 0
        result.append({
            "material_id": mat_id,
            "material_name": mat.name,
            "unit": mat.unit,
            "required_qty": round(req_qty, 3),
            "current_stock": round(stock, 3),
            "sufficient": stock >= req_qty,
            "shortage": round(stock - req_qty, 3),
        })

    return sorted(result, key=lambda x: x["material_name"])

@app.get("/api/production-plans")
def list_production_plans(
    year: Optional[int] = None,
    month: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db)
):
    q = db.query(ProductionPlan).filter(ProductionPlan.user_id == uid()).order_by(ProductionPlan.planned_date)
    if year and month:
        from calendar import monthrange
        _, last_day = monthrange(year, month)
        q = q.filter(
            ProductionPlan.planned_date >= datetime(year, month, 1),
            ProductionPlan.planned_date <= datetime(year, month, last_day, 23, 59, 59),
        )
    elif date_from and date_to:
        q = q.filter(
            ProductionPlan.planned_date >= datetime.fromisoformat(date_from),
            ProductionPlan.planned_date <= datetime.fromisoformat(date_to + "T23:59:59"),
        )
    result = []
    for p in q.all():
        # BOM 부족 여부 계산
        has_shortage = False
        try:
            mat_reqs = explode_bom(p.product_id, p.planned_quantity, db)
            for mat_id, req_qty in mat_reqs.items():
                mat = db.query(Material).filter(Material.id == mat_id, Material.user_id == uid()).first()
                if mat and (mat.current_stock or 0) < req_qty:
                    has_shortage = True
                    break
        except Exception:
            pass
        result.append({
            "id": p.id,
            "product_id": p.product_id,
            "product_name": p.product.name if p.product else "",
            "product_unit": p.product.unit if p.product else "",
            "planned_date": p.planned_date.strftime("%Y-%m-%d"),
            "planned_quantity": p.planned_quantity,
            "note": p.note,
            "status": p.status,
            "has_shortage": has_shortage,
        })
    return result

@app.post("/api/production-plans", status_code=201)
def create_production_plan(data: ProductionPlanCreate, db: Session = Depends(get_db)):
    if not db.query(FinishedProduct).filter(FinishedProduct.id == data.product_id, FinishedProduct.user_id == uid()).first():
        raise HTTPException(404, "Product not found")
    plan = ProductionPlan(
        product_id=data.product_id,
        planned_date=datetime.fromisoformat(data.planned_date),
        planned_quantity=data.planned_quantity,
        note=data.note,
        status=data.status,
        user_id=uid(),
    )
    db.add(plan); db.commit(); db.refresh(plan)
    return {"id": plan.id}

@app.put("/api/production-plans/{pid}")
def update_production_plan(pid: int, data: ProductionPlanCreate, db: Session = Depends(get_db)):
    plan = db.query(ProductionPlan).filter(ProductionPlan.id == pid, ProductionPlan.user_id == uid()).first()
    if not plan:
        raise HTTPException(404, "Not found")
    plan.product_id = data.product_id
    plan.planned_date = datetime.fromisoformat(data.planned_date)
    plan.planned_quantity = data.planned_quantity
    plan.note = data.note
    plan.status = data.status
    db.commit()
    return {"ok": True}

@app.delete("/api/production-plans/{pid}")
def delete_production_plan(pid: int, db: Session = Depends(get_db)):
    plan = db.query(ProductionPlan).filter(ProductionPlan.id == pid, ProductionPlan.user_id == uid()).first()
    if not plan:
        raise HTTPException(404, "Not found")
    db.delete(plan); db.commit()
    return {"ok": True}


# ─────────────────────────────────────────
# 출고 관리
# ─────────────────────────────────────────
def _shipment_dict(s: Shipment) -> dict:
    return {
        "id": s.id,
        "shipment_number": s.shipment_number,
        "finished_product_id": s.finished_product_id,
        "product_name": s.finished_product.name if s.finished_product else "",
        "product_unit": s.finished_product.unit if s.finished_product else "",
        "customer_id": s.customer_id,
        "customer_name": s.customer.name if s.customer else "",
        "production_id": s.production_id,
        "lot_number": s.lot_number or "",
        "quantity": s.quantity,
        "unit_price": s.unit_price,
        "total_amount": s.total_amount,
        "delivery_date": s.delivery_date.strftime("%Y-%m-%d") if s.delivery_date else None,
        "note": s.note,
        "status": s.status,
        "created_at": s.created_at.strftime("%Y-%m-%d %H:%M") if s.created_at else "",
    }


@app.get("/api/shipments")
def list_shipments(
    limit: int = 200,
    product_id: Optional[int] = None,
    customer_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(Shipment).filter(Shipment.user_id == uid()).order_by(desc(Shipment.delivery_date))
    if product_id:
        q = q.filter(Shipment.finished_product_id == product_id)
    if customer_id:
        q = q.filter(Shipment.customer_id == customer_id)
    if date_from:
        q = q.filter(Shipment.delivery_date >= datetime.fromisoformat(date_from))
    if date_to:
        q = q.filter(Shipment.delivery_date < datetime.fromisoformat(date_to) + timedelta(days=1))
    return [_shipment_dict(s) for s in q.limit(limit).all()]


@app.post("/api/shipments", status_code=201)
def create_shipment(data: ShipmentCreate, db: Session = Depends(get_db)):
    product = db.query(FinishedProduct).filter(
        FinishedProduct.id == data.finished_product_id, FinishedProduct.user_id == uid()
    ).first()
    if not product:
        raise HTTPException(404, "Product not found")

    now = datetime.now()
    shipment_number = f"SH-{now.strftime('%Y%m%d')}-{secrets.token_hex(2).upper()}"

    lot = data.lot_number
    if not lot and data.production_id:
        prod_row = db.query(Production).filter(
            Production.id == data.production_id, Production.user_id == uid()
        ).first()
        if prod_row:
            lot = prod_row.lot_number
    if not lot:
        lot = f"SHLOT-{now.strftime('%Y%m%d')}-{secrets.token_hex(2).upper()}"

    unit_price = data.unit_price if data.unit_price is not None else product.unit_price
    total = round(data.quantity * unit_price, 2) if unit_price else None

    shipment = Shipment(
        shipment_number=shipment_number,
        finished_product_id=data.finished_product_id,
        customer_id=data.customer_id,
        production_id=data.production_id,
        lot_number=lot,
        quantity=data.quantity,
        unit_price=unit_price,
        total_amount=total,
        delivery_date=datetime.fromisoformat(data.delivery_date) if data.delivery_date else now,
        note=data.note,
        status="confirmed",
        user_id=uid(),
    )
    db.add(shipment)
    product.current_stock = max(0, (product.current_stock or 0) - data.quantity)
    db.commit()
    db.refresh(shipment)
    return {"id": shipment.id, "shipment_number": shipment.shipment_number}


@app.put("/api/shipments/{shipment_id}")
def update_shipment(shipment_id: int, data: ShipmentUpdate, db: Session = Depends(get_db)):
    s = db.query(Shipment).filter(Shipment.id == shipment_id, Shipment.user_id == uid()).first()
    if not s:
        raise HTTPException(404, "Not found")

    old_pid = s.finished_product_id
    old_qty = s.quantity
    new_pid = data.finished_product_id if data.finished_product_id is not None else old_pid
    new_qty = data.quantity if data.quantity is not None else old_qty

    # 기존 재고 복원
    old_fp = db.query(FinishedProduct).filter(FinishedProduct.id == old_pid, FinishedProduct.user_id == uid()).first()
    if old_fp:
        old_fp.current_stock = (old_fp.current_stock or 0) + old_qty

    # 새 제품 재고 차감
    new_fp = db.query(FinishedProduct).filter(FinishedProduct.id == new_pid, FinishedProduct.user_id == uid()).first()
    if not new_fp:
        raise HTTPException(404, "Product not found")
    new_fp.current_stock = max(0, (new_fp.current_stock or 0) - new_qty)

    new_price = data.unit_price if data.unit_price is not None else s.unit_price
    new_total = round(new_qty * new_price, 2) if new_price else None

    s.finished_product_id = new_pid
    s.quantity = new_qty
    s.unit_price = new_price
    s.total_amount = new_total
    if data.customer_id is not None:
        s.customer_id = data.customer_id or None
    if data.lot_number is not None:
        s.lot_number = data.lot_number
    if data.delivery_date is not None:
        s.delivery_date = datetime.fromisoformat(data.delivery_date)
    if data.note is not None:
        s.note = data.note

    db.commit()
    return _shipment_dict(s)


@app.delete("/api/shipments/{shipment_id}")
def delete_shipment(shipment_id: int, db: Session = Depends(get_db)):
    s = db.query(Shipment).filter(Shipment.id == shipment_id, Shipment.user_id == uid()).first()
    if not s:
        raise HTTPException(404, "Not found")
    product = db.query(FinishedProduct).filter(
        FinishedProduct.id == s.finished_product_id, FinishedProduct.user_id == uid()
    ).first()
    if product:
        product.current_stock = (product.current_stock or 0) + s.quantity
    db.delete(s)
    db.commit()
    return {"ok": True}


@app.get("/api/shipments/{shipment_id}/statement")
def get_statement(shipment_id: int, db: Session = Depends(get_db)):
    s = db.query(Shipment).filter(Shipment.id == shipment_id, Shipment.user_id == uid()).first()
    if not s:
        raise HTTPException(404, "Not found")
    user = db.query(TenantUser).filter(TenantUser.id == uid()).first()
    return {
        "shipment_number": s.shipment_number,
        "delivery_date": s.delivery_date.strftime("%Y-%m-%d") if s.delivery_date else "",
        "supplier": {
            "company_name": user.company_name if user else "",
            "business_number": user.business_number if user else "",
            "contact": user.contact if user else "",
            "seal_image": user.seal_image if user else None,
        },
        "customer": {
            "name": s.customer.name if s.customer else "",
            "business_number": s.customer.business_number if s.customer else "",
            "address": s.customer.address if s.customer else "",
            "contact": s.customer.contact if s.customer else "",
            "contact_person": s.customer.contact_person if s.customer else "",
        },
        "items": [{
            "product_name": s.finished_product.name if s.finished_product else "",
            "unit": s.finished_product.unit if s.finished_product else "",
            "lot_number": s.lot_number or "",
            "quantity": s.quantity,
            "unit_price": s.unit_price,
            "total_amount": s.total_amount,
        }],
        "total_amount": s.total_amount,
        "note": s.note or "",
    }


# ─────────────────────────────────────────
# 매출 대시보드
# ─────────────────────────────────────────
@app.get("/api/sales/dashboard")
def sales_dashboard(db: Session = Depends(get_db)):
    _uid = uid()
    today = datetime.now().date()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    def _amt(s):
        return s.total_amount or (s.quantity * (s.unit_price or 0))

    today_rows = db.query(Shipment).filter(
        Shipment.user_id == _uid, func.date(Shipment.delivery_date) == today
    ).all()
    week_rows = db.query(Shipment).filter(
        Shipment.user_id == _uid,
        Shipment.delivery_date >= datetime.combine(week_start, datetime.min.time()),
    ).all()
    month_rows = db.query(Shipment).filter(
        Shipment.user_id == _uid,
        Shipment.delivery_date >= datetime.combine(month_start, datetime.min.time()),
    ).all()

    # 제품별 매출 (당월)
    product_sales: dict = {}
    for s in month_rows:
        name = s.finished_product.name if s.finished_product else "기타"
        product_sales[name] = product_sales.get(name, 0) + _amt(s)
    product_chart = [{"name": k, "amount": round(v, 0)}
                     for k, v in sorted(product_sales.items(), key=lambda x: -x[1])]

    # 거래처별 매출 (당월)
    customer_sales: dict = {}
    for s in month_rows:
        name = s.customer.name if s.customer else "미지정"
        customer_sales[name] = customer_sales.get(name, 0) + _amt(s)
    customer_chart = [{"name": k, "amount": round(v, 0), "count": sum(
        1 for ss in month_rows if (ss.customer.name if ss.customer else "미지정") == k
    )} for k, v in sorted(customer_sales.items(), key=lambda x: -x[1])]

    # 최근 7일 일별 매출
    day_chart = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        day_rows = db.query(Shipment).filter(
            Shipment.user_id == _uid, func.date(Shipment.delivery_date) == d
        ).all()
        day_chart.append({
            "date": d.strftime("%m/%d"),
            "amount": round(sum(_amt(s) for s in day_rows), 0),
        })

    return {
        "today_sales": round(sum(_amt(s) for s in today_rows), 0),
        "today_count": len(today_rows),
        "week_sales": round(sum(_amt(s) for s in week_rows), 0),
        "week_count": len(week_rows),
        "month_sales": round(sum(_amt(s) for s in month_rows), 0),
        "month_count": len(month_rows),
        "product_chart": product_chart,
        "customer_chart": customer_chart,
        "day_chart": day_chart,
    }
